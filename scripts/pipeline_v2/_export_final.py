"""Export final Excel with all combined data for 27 test SKUs."""
from __future__ import annotations

import json
import re
import sys
import io
from pathlib import Path
from urllib.parse import urlparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
EV_DIR = ROOT / "downloads" / "evidence"

tier_results = json.loads(
    (ROOT / "downloads/staging/tier_collector_output/tier_collector_results.json").read_text(encoding="utf-8"))
ds_results = json.loads(
    (ROOT / "downloads/staging/tier_collector_output/datasheet_extracted.json").read_text(encoding="utf-8"))
audit_file = ROOT / "downloads" / "photos_ai_v2" / "_quality_audit.json"
audit = json.loads(audit_file.read_text(encoding="utf-8")) if audit_file.exists() else {}

TEST_PNS = [
    "109411", "EASY", "CM010610", "00020211", "010130.10", "1000106",
    "1011893-RU", "1050000000", "2208WFPT", "CF274A", "CWSS-RB-S8",
    "7508001857", "EVCS-HSB", "36022-RU", "3240197-RU", "2CDG110146R0011",
    "7910180000", "1006186", "1006187", "027913.10", "2SM-3.0-SCU-SCU-1",
    "CAB-010-SC-SM", "36299-RU", "2CDG110177R0011", "36024-RU",
    "1011894-RU", "7508001858",
]

HEADERS = [
    "PN", "Brand", "Seed Name (Excel)", "Title (best)",
    "Price (evidence)", "Currency", "Price (new)", "New Currency",
    "EAN", "EAN Source",
    "Description RU (short)", "Photo Status", "Photo File",
    "Weight (g)", "Dimensions (mm)",
    "Specs Count", "Top Specs",
    "Datasheet", "Datasheet Source",
    "Category (DR)", "Identity Level",
    "Data Sources",
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pipeline V2 (27 SKU)"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, pn in enumerate(TEST_PNS, 2):
        ef = EV_DIR / f"evidence_{pn}.json"
        d = json.loads(ef.read_text(encoding="utf-8"))
        si = d.get("structured_identity") or {}
        norm = d.get("normalized") or {}
        dr = d.get("deep_research") or {}
        content = d.get("content") or {}

        brand = d.get("brand", "") or si.get("confirmed_manufacturer", "")
        subbrand = d.get("subbrand", "")
        real_brand = subbrand or brand
        seed_name = content.get("seed_name", "") or d.get("name", "")

        # Title
        title = dr.get("title_ru", "") or d.get("assembled_title", "") or seed_name
        if title.lower() == "unknown":
            title = d.get("assembled_title", "") or seed_name

        # Evidence price
        ev_price = norm.get("best_price", "")
        ev_currency = norm.get("best_price_currency", "")

        # Playwright price
        new_price = ""
        new_currency = ""
        pw_data = None
        for r in tier_results.get(pn, []):
            if r.get("is_correct_product") and r.get("price"):
                new_price = r["price"]
                new_currency = r.get("currency", "")
                pw_data = r
                break

        # EAN — check ALL sources
        ean = ""
        ean_source = ""
        ds_data = ds_results.get(pn)
        # Datasheet EAN first (most authoritative)
        if ds_data and ds_data.get("ean"):
            ean_val = str(ds_data["ean"])
            if ean_val and ean_val not in ("Not provided", "Not specified", "", "-", "None"):
                ean = ean_val
                ean_source = "datasheet"
        # Playwright EAN second
        if not ean and pw_data and pw_data.get("ean"):
            ean_val = str(pw_data["ean"])
            if ean_val and ean_val not in ("Not provided", "Not specified", "", "-"):
                ean = ean_val
                ean_source = pw_data.get("_source_domain", "playwright")

        # Description
        desc = (norm.get("best_description", "") or "")[:300]

        # Photo
        audit_result = audit.get(pn, {})
        is_wrong = (audit_result.get("primary_issue") == "wrong_product"
                    or "wrong_product" in audit_result.get("issues", []))
        pn_safe = pn.replace("/", "_").replace(" ", "_")
        has_photo_file = (ROOT / f"downloads/photos_ai_v2/{pn_safe}.jpg").exists()

        # Check for datasheet photos
        ds_photos = list((ROOT / "downloads" / "staging" / "pipeline_v2_export" / "photos").glob(f"{pn_safe}_ds_*"))

        if is_wrong:
            photo_status = "WRONG_PRODUCT"
            if ds_photos:
                photo_file = ds_photos[0].name
                photo_status = "FROM_DATASHEET"
            else:
                photo_file = ""
        elif has_photo_file and audit_result:
            photo_status = audit_result.get("primary_issue", "good").upper()
            photo_file = f"{pn_safe}.jpg"
        elif has_photo_file:
            photo_status = "NOT_AUDITED"
            photo_file = f"{pn_safe}.jpg"
        elif ds_photos:
            photo_status = "FROM_DATASHEET"
            photo_file = ds_photos[0].name
        else:
            photo_status = "NO_PHOTO"
            photo_file = ""

        # Weight/Dimensions from datasheet
        weight = ""
        dims = ""
        if ds_data:
            w = ds_data.get("weight_g", "")
            if isinstance(w, dict):
                weight = w.get("net", "") or w.get("product_net", "")
            elif w and str(w) not in ("Not specified", "Not provided", "?", ""):
                weight = str(w)

            dm = ds_data.get("dimensions_mm", "")
            if isinstance(dm, dict):
                parts = []
                for k in ["depth_length", "length", "width", "height"]:
                    if dm.get(k):
                        parts.append(str(dm[k]))
                dims = " x ".join(parts)
            elif dm and str(dm) not in ("Not specified", "Not provided", "?", ""):
                dims = str(dm)

        # Specs
        all_specs = {}
        if pw_data and pw_data.get("specs"):
            all_specs.update(pw_data["specs"])
        if ds_data and ds_data.get("specs"):
            all_specs.update(ds_data["specs"])
        specs_count = len(all_specs)
        top_specs = "; ".join(f"{k}={v}" for k, v in list(all_specs.items())[:5])

        # Datasheet
        has_ds = (ROOT / f"downloads/datasheets_v2/{pn_safe}.pdf").exists()
        ds_source = ""
        if has_ds:
            raw_ev = json.dumps(d)
            pdf_urls = re.findall(r"https?://[^\s\"<>]+\.pdf", raw_ev, re.IGNORECASE)
            if pdf_urls:
                ds_source = urlparse(pdf_urls[0]).netloc.replace("www.", "")

        # Category / identity
        category = d.get("dr_category", "")
        identity_level = d.get("identity_level", "")

        # Sources
        sources = []
        if has_ds:
            sources.append("datasheet")
        if pw_data:
            sources.append(pw_data.get("_source_domain", ""))
        if ev_price:
            sources.append("evidence")
        sources_str = ", ".join(s for s in sources if s)

        values = [
            pn, real_brand, seed_name, title,
            ev_price, ev_currency, new_price, new_currency,
            ean, ean_source,
            desc, photo_status, photo_file,
            weight, dims,
            specs_count, top_specs,
            "YES" if has_ds else "NO", ds_source,
            category, identity_level,
            sources_str,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Auto-width
    for col in range(1, len(HEADERS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(TEST_PNS) + 2)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 60)

    out = ROOT / "downloads" / "staging" / "pipeline_v2_export" / "catalog_v2_test_29sku.xlsx"
    wb.save(str(out))
    print(f"Excel: {out}")

    # Summary
    counts = {
        "new_price": 0, "ean": 0, "datasheet": 0,
        "specs": 0, "valid_photo": 0, "ev_price": 0,
    }
    for r in range(2, len(TEST_PNS) + 2):
        if ws.cell(row=r, column=7).value:
            counts["new_price"] += 1
        if ws.cell(row=r, column=9).value:
            counts["ean"] += 1
        if ws.cell(row=r, column=18).value == "YES":
            counts["datasheet"] += 1
        if (ws.cell(row=r, column=16).value or 0) > 0:
            counts["specs"] += 1
        if ws.cell(row=r, column=13).value:
            counts["valid_photo"] += 1
        if ws.cell(row=r, column=5).value:
            counts["ev_price"] += 1

    print("\n27 SKUs summary:")
    print(f"  Price (evidence):     {counts['ev_price']}")
    print(f"  Price (new/Playwright): {counts['new_price']}")
    print(f"  EAN found:            {counts['ean']}")
    print(f"  Datasheet downloaded: {counts['datasheet']}")
    print(f"  Specs extracted:      {counts['specs']}")
    print(f"  Valid photo:          {counts['valid_photo']}")


if __name__ == "__main__":
    main()
