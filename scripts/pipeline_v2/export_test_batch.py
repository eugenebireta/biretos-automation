"""Export test batch: photos folder + Excel for 29 test SKUs.

Creates:
1. downloads/staging/pipeline_v2_export/photos/ — local photos
2. downloads/staging/pipeline_v2_export/catalog_v2_test.xlsx — filled Excel
"""
from __future__ import annotations

import json
import shutil
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

ROOT = Path(__file__).resolve().parent.parent.parent
EXPORT_DIR = ROOT / "downloads" / "staging" / "pipeline_v2_export"
PHOTOS_DIR = EXPORT_DIR / "photos"
CANONICAL_FILE = ROOT / "downloads" / "staging" / "pipeline_v2_output" / "canonical_products.json"
EV_DIR = ROOT / "downloads" / "evidence"

# Photo source directories (priority order)
PHOTO_DIRS = [
    ROOT / "downloads" / "photos_ai_v2_good",
    ROOT / "downloads" / "photos_ai_v2",
    ROOT / "downloads" / "photos_best",
    ROOT / "downloads" / "photos",
]

TEST_PNS = [
    "109411", "EASY", "CM010610", "272369098",
    "00020211", "010130.10", "1000106", "1011893-RU",
    "1050000000", "2208WFPT", "CF274A", "CWSS-RB-S8",
    "7508001857", "D-71570", "EVCS-HSB", "36022-RU",
    "3240197-RU", "2CDG110146R0011", "7910180000",
    "1006186", "1006187", "027913.10", "2SM-3.0-SCU-SCU-1",
    "CAB-010-SC-SM", "36299-RU", "2CDG110177R0011",
    "36024-RU", "1011894-RU", "7508001858",
]


def find_local_photo(pn: str) -> Path | None:
    """Find best local photo for PN."""
    pn_safe = pn.replace("/", "_").replace(" ", "_").replace("\\", "_")
    for d in PHOTO_DIRS:
        p = d / f"{pn_safe}.jpg"
        if p.exists():
            return p
    return None


def main():
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    PHOTOS_DIR.mkdir(parents=True, exist_ok=True)

    # Load canonical products
    products = json.loads(CANONICAL_FILE.read_text(encoding="utf-8"))
    product_map = {p["identity"]["pn"]: p for p in products}

    # Load weak PNs
    weak_file = ROOT / "downloads" / "staging" / "pipeline_v2_output" / "weak_pns.json"
    weak_pns = {w["pn"] for w in json.loads(weak_file.read_text(encoding="utf-8"))}

    print(f"Exporting {len(TEST_PNS)} test SKUs")
    print(f"Output: {EXPORT_DIR}")
    print()

    # Collect rows for Excel
    rows = []
    photos_copied = 0
    photos_missing = 0

    for pn in TEST_PNS:
        # Load evidence for extra fields
        ef = EV_DIR / f"evidence_{pn}.json"
        ev = json.loads(ef.read_text(encoding="utf-8")) if ef.exists() else {}
        content = ev.get("content") or {}
        norm = ev.get("normalized") or {}
        ev.get("structured_identity") or {}
        ev.get("deep_research") or {}

        if pn in product_map:
            p = product_map[pn]
            verdict = "CONFIRMED"
            brand = p["identity"]["brand"]
            pn_val = p["identity"]["pn"]
            product_type = p["identity"].get("product_type") or ""
            series = p["identity"].get("series") or ""
            ean = p["identity"].get("ean") or ""
            # Title: from canonical (now uses assembled_title/seed_name/dr_title)
            title_ru = p.get("title_ru") or ""
            # If still looks like bare "Brand PN", use seed_name
            if title_ru and title_ru == f"{brand} {pn_val}":
                title_ru = p["identity"].get("seed_name", "") or title_ru
            best_price = p.get("best_price")
            best_price_currency = p.get("best_price_currency") or ""
            best_description = p.get("best_description_ru") or ""
            best_photo_url = p.get("best_photo_url") or ""
            insales_ready = p["readiness"]["insales"]
            specs = p.get("specs") or {}
            weight_g = specs.get("weight_g")
            dimensions = ""
            if specs.get("length_mm") and specs.get("width_mm"):
                dimensions = f"{specs['length_mm']}x{specs['width_mm']}"
                if specs.get("height_mm"):
                    dimensions += f"x{specs['height_mm']}"
                dimensions += " mm"
            material = specs.get("material") or ""
            color = specs.get("color_canonical") or ""
            ip_rating = specs.get("ip_rating") or ""

            # Trusted sources
            trusted = p.get("trusted_sources") or []
            trusted_str = "; ".join(f"{s['domain']}({','.join(s['has'])})" for s in trusted[:5])

            bound_count = (p.get("evidence_stats") or {}).get("bound_count", 0)
        elif pn in weak_pns:
            verdict = "WEAK"
            brand = ev.get("brand", "")
            pn_val = pn
            product_type = ""
            series = ""
            ean = ""
            title_ru = ev.get("assembled_title", "")
            best_price = norm.get("best_price")
            best_price_currency = norm.get("best_price_currency", "")
            best_description = norm.get("best_description", "")
            best_photo_url = norm.get("best_photo_url", "")
            insales_ready = "WEAK_IDENTITY"
            weight_g = None
            dimensions = ""
            material = ""
            color = ""
            ip_rating = ""
            trusted_str = ""
            bound_count = 0
            specs = {}
        else:
            verdict = "REJECTED"
            brand = ev.get("brand", "")
            pn_val = pn
            product_type = ""
            series = ""
            ean = ""
            title_ru = ev.get("assembled_title", "")
            best_price = None
            best_price_currency = ""
            best_description = ""
            best_photo_url = ""
            insales_ready = "REJECTED"
            weight_g = None
            dimensions = ""
            material = ""
            color = ""
            ip_rating = ""
            trusted_str = ""
            bound_count = 0
            specs = {}

        # Copy photo (only if passed product-aware audit)
        audit_file = ROOT / "downloads" / "photos_ai_v2" / "_quality_audit.json"
        if not hasattr(main, '_audit_data'):
            main._audit_data = json.loads(audit_file.read_text(encoding="utf-8")) if audit_file.exists() else {}
        audit_result = main._audit_data.get(pn, {})
        is_wrong = audit_result.get("primary_issue") == "wrong_product" or "wrong_product" in audit_result.get("issues", [])

        local_photo = find_local_photo(pn)
        photo_filename = ""
        if local_photo and not is_wrong:
            pn_safe = pn.replace("/", "_").replace(" ", "_")
            photo_filename = f"{pn_safe}.jpg"
            dest = PHOTOS_DIR / photo_filename
            shutil.copy2(local_photo, dest)
            photos_copied += 1
        elif is_wrong:
            photos_missing += 1  # wrong product photo = same as missing
        else:
            photos_missing += 1

        # Excel data from evidence
        our_price_raw = ev.get("our_price_raw", "")
        seed_name = content.get("seed_name", "") or ev.get("name", "")
        dr_category = ev.get("dr_category", "")
        identity_level = ev.get("identity_level", "")

        rows.append({
            "PN": pn_val,
            "Brand": brand,
            "Verdict": verdict,
            "Identity Level": identity_level,
            "Product Type": product_type,
            "Series": series,
            "EAN": ean,
            "Title RU": title_ru,
            "Seed Name (Excel)": seed_name,
            "Our Price (RUB)": our_price_raw,
            "Best Price": best_price,
            "Best Price Currency": best_price_currency,
            "Description RU": best_description[:500] if best_description else "",
            "Photo File": photo_filename,
            "Photo URL": best_photo_url,
            "Category (DR)": dr_category,
            "Weight (g)": weight_g or "",
            "Dimensions": dimensions,
            "Material": material,
            "Color": color,
            "IP Rating": ip_rating,
            "Trusted Sources": trusted_str,
            "Bound Evidence Count": bound_count,
            "InSales Ready": insales_ready,
        })

    # Write Excel
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pipeline V2 Test"

        # Header
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data
        for row_idx, row in enumerate(rows, 2):
            for col, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=row[h])

        # Auto-width
        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(rows) + 2))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 50)

        excel_path = EXPORT_DIR / "catalog_v2_test_29sku.xlsx"
        wb.save(str(excel_path))
        print(f"Excel: {excel_path}")
    except ImportError:
        # Fallback to CSV
        import csv
        csv_path = EXPORT_DIR / "catalog_v2_test_29sku.csv"
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys(), delimiter=";")
            writer.writeheader()
            writer.writerows(rows)
        print(f"CSV (no openpyxl): {csv_path}")

    print(f"Photos: {PHOTOS_DIR} ({photos_copied} copied, {photos_missing} missing)")
    print()
    print("Summary:")
    confirmed = sum(1 for r in rows if r["Verdict"] == "CONFIRMED")
    ready = sum(1 for r in rows if r["InSales Ready"] == "READY")
    with_photo = sum(1 for r in rows if r["Photo File"])
    with_price = sum(1 for r in rows if r["Best Price"])
    print(f"  CONFIRMED: {confirmed}/29")
    print(f"  InSales READY: {ready}/29")
    print(f"  With local photo: {with_photo}/29")
    print(f"  With price: {with_price}/29")


if __name__ == "__main__":
    main()
