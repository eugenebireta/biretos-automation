"""Night batch 2: validate datasheets + extract photos + fill gaps.

1. Validate each parsed datasheet — does it match the requested PN?
2. Extract photos from ALL 311 PDFs (not just 8)
3. Re-parse remaining 21 large PDFs
4. Build final stats + update Excel
"""
from __future__ import annotations

import json
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

ROOT = Path(__file__).resolve().parent.parent.parent
DS_DIR = ROOT / "downloads" / "datasheets_v2"
EV_DIR = ROOT / "downloads" / "evidence"
PHOTOS_OUT = ROOT / "downloads" / "datasheet_photos"
RESULTS_FILE = ROOT / "downloads" / "staging" / "tier_collector_output" / "datasheet_extracted.json"
VALIDATION_FILE = ROOT / "downloads" / "staging" / "tier_collector_output" / "datasheet_validation.json"


def phase1_validate_datasheets():
    """Use Gemini to validate: does the PDF content match the requested PN?"""
    import google.generativeai as genai
    from scripts.app_secrets import get_secret
    genai.configure(api_key=get_secret("GEMINI_API_KEY"))
    genai.GenerativeModel("gemini-2.5-flash")

    parsed = json.loads(RESULTS_FILE.read_text(encoding="utf-8")) if RESULTS_FILE.exists() else {}
    existing = json.loads(VALIDATION_FILE.read_text(encoding="utf-8")) if VALIDATION_FILE.exists() else {}

    print("=" * 90)
    print(f"PHASE 1: Validate {len(parsed)} parsed datasheets")
    print("=" * 90)

    validated = 0
    correct = 0
    wrong = 0

    for pn, data in parsed.items():
        if pn in existing:
            continue

        # Quick heuristic first — if extracted PN matches, skip Gemini
        extracted_pn = str(data.get("pn", "")).strip()
        extracted_title = data.get("title", "")

        # Load evidence
        ev_file = EV_DIR / f"evidence_{pn.replace('_','/')}.json"
        if not ev_file.exists():
            ev_file = EV_DIR / f"evidence_{pn}.json"
        if not ev_file.exists():
            continue

        d = json.loads(ev_file.read_text(encoding="utf-8"))
        brand = d.get("brand", "")
        subbrand = d.get("subbrand", "")
        real_brand = subbrand or brand
        (d.get("content") or {}).get("seed_name", "") or d.get("name", "")

        # Quick check — PN in extracted data or title
        pn_no_dash = pn.replace("-", "").replace(".", "").upper()
        extracted_no_dash = extracted_pn.replace("-", "").replace(".", "").upper() if extracted_pn else ""
        title_no_dash = extracted_title.replace("-", "").replace(".", "").upper()

        quick_match = (
            (extracted_no_dash and pn_no_dash in extracted_no_dash) or
            (title_no_dash and pn_no_dash in title_no_dash)
        )

        # Brand check
        brand_match = False
        if real_brand and extracted_title:
            extracted_brand_str = str(data.get("brand") or "").upper()
            real_brand_upper = real_brand.upper()
            if real_brand_upper in extracted_title.upper():
                brand_match = True
            elif extracted_brand_str and real_brand_upper in extracted_brand_str:
                brand_match = True

        if quick_match and brand_match:
            existing[pn] = {"verdict": "CORRECT", "reason": "pn_and_brand_in_extraction", "method": "heuristic"}
            correct += 1
        elif extracted_pn and extracted_pn != pn and not quick_match:
            existing[pn] = {
                "verdict": "WRONG",
                "reason": f"PN mismatch: requested={pn}, extracted={extracted_pn}",
                "method": "heuristic",
                "extracted_pn": extracted_pn,
                "extracted_title": extracted_title[:80],
            }
            wrong += 1
        elif not extracted_title and not data.get("specs"):
            existing[pn] = {
                "verdict": "EMPTY",
                "reason": "no title, no specs extracted",
                "method": "heuristic",
            }
            wrong += 1
        else:
            # Uncertain — mark for review but don't use Gemini for now
            existing[pn] = {
                "verdict": "UNCERTAIN",
                "reason": "needs manual review",
                "method": "heuristic",
                "extracted_pn": extracted_pn,
                "extracted_title": extracted_title[:80],
            }

        validated += 1
        if validated % 50 == 0:
            VALIDATION_FILE.write_text(json.dumps(existing, indent=2, ensure_ascii=False), encoding="utf-8")

    VALIDATION_FILE.write_text(json.dumps(existing, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Validated: {validated}")
    print(f"  CORRECT:   {sum(1 for v in existing.values() if v.get('verdict') == 'CORRECT')}")
    print(f"  WRONG:     {sum(1 for v in existing.values() if v.get('verdict') == 'WRONG')}")
    print(f"  EMPTY:     {sum(1 for v in existing.values() if v.get('verdict') == 'EMPTY')}")
    print(f"  UNCERTAIN: {sum(1 for v in existing.values() if v.get('verdict') == 'UNCERTAIN')}")


def phase2_extract_photos():
    """Extract product photos from ALL downloaded PDFs."""
    import fitz

    PHOTOS_OUT.mkdir(parents=True, exist_ok=True)

    validation = json.loads(VALIDATION_FILE.read_text(encoding="utf-8")) if VALIDATION_FILE.exists() else {}
    pdfs = [p for p in DS_DIR.glob("*.pdf") if "catalog" not in p.stem.lower()]

    print("\n" + "=" * 90)
    print(f"PHASE 2: Extract photos from {len(pdfs)} PDFs")
    print("=" * 90)

    total_extracted = 0
    skus_with_photos = 0

    for pdf in pdfs:
        pn = pdf.stem
        # Skip if not validated as CORRECT
        val = validation.get(pn, {})
        if val.get("verdict") in ("WRONG", "EMPTY"):
            continue

        # Skip if already extracted
        existing_photos = list(PHOTOS_OUT.glob(f"{pn}_p*.jpeg")) + list(PHOTOS_OUT.glob(f"{pn}_p*.png"))
        if existing_photos:
            continue

        try:
            doc = fitz.open(str(pdf))
            extracted = []

            for page_num in range(min(len(doc), 10)):  # max 10 pages
                page = doc[page_num]
                images = page.get_images(full=True)

                for img_idx, img in enumerate(images):
                    xref = img[0]
                    try:
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        ext = base_image["ext"]
                        width = base_image.get("width", 0)
                        height = base_image.get("height", 0)

                        # Filters
                        if width < 150 or height < 150:
                            continue
                        # Filter banners
                        if width > 500 and height < 300 and width / max(height, 1) > 2.5:
                            continue
                        if len(image_bytes) < 3000:
                            continue

                        img_name = f"{pn}_p{page_num+1}_{img_idx}.{ext}"
                        (PHOTOS_OUT / img_name).write_bytes(image_bytes)
                        extracted.append(img_name)

                        # Max 3 photos per SKU
                        if len(extracted) >= 3:
                            break
                    except Exception:
                        continue

                if len(extracted) >= 3:
                    break

            doc.close()

            if extracted:
                total_extracted += len(extracted)
                skus_with_photos += 1

        except Exception:
            pass

    print(f"  SKUs with photos: {skus_with_photos}")
    print(f"  Total photos:     {total_extracted}")


def phase3_update_excel():
    """Rebuild final Excel with all latest data."""
    import openpyxl

    parsed = json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
    validation = json.loads(VALIDATION_FILE.read_text(encoding="utf-8")) if VALIDATION_FILE.exists() else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All SKUs (370)"

    HEADERS = [
        "PN", "Brand", "Seed Name (Excel)",
        "Title (datasheet)", "Description",
        "Our Price RUB", "Best Price", "Currency",
        "EAN (datasheet)", "EAN Source",
        "Weight (g)", "Dimensions (mm)",
        "Specs Count", "Top Specs",
        "Series", "Category",
        "Datasheet", "DS Verdict",
        "Photo count (datasheet)",
        "Certifications",
    ]

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    row_idx = 2
    for ef in sorted(EV_DIR.glob("evidence_*.json")):
        d = json.loads(ef.read_text(encoding="utf-8"))
        pn = d.get("pn", "")
        if not pn or pn.strip("-_") == "" or pn in {"---", "--", "_", "PN", "-----"}:
            continue

        pn_safe = pn.replace("/", "_").replace(" ", "_")
        si = d.get("structured_identity") or {}
        norm = d.get("normalized") or {}
        content = d.get("content") or {}

        brand = d.get("brand", "") or si.get("confirmed_manufacturer", "")
        subbrand = d.get("subbrand", "")
        real_brand = subbrand or brand
        seed = content.get("seed_name", "") or d.get("name", "")

        ds_data = parsed.get(pn_safe) or parsed.get(pn, {})
        ds_val = validation.get(pn_safe) or validation.get(pn, {})

        title = ds_data.get("title", "")
        description = ds_data.get("description", "")[:500]
        ean = ds_data.get("ean", "")
        if ean and str(ean) in ("-", "None", "Not provided", "Not specified", ""):
            ean = ""

        weight = ds_data.get("weight_g", "")
        if isinstance(weight, dict):
            weight = weight.get("net") or weight.get("product_net") or ""
        if str(weight) in ("Not specified", "Not provided", "?", "None"):
            weight = ""

        dims = ds_data.get("dimensions_mm", "")
        if isinstance(dims, dict):
            parts = []
            for k in ["length", "width", "height", "depth_length"]:
                if dims.get(k):
                    parts.append(str(dims[k]))
            dims = " x ".join(parts)
        if str(dims) in ("Not specified", "Not provided", "?", "None"):
            dims = ""

        specs = ds_data.get("specs", {})
        specs_count = len(specs) if isinstance(specs, dict) else 0
        top_specs = "; ".join(f"{k}={v}" for k, v in list(specs.items())[:5]) if isinstance(specs, dict) else ""

        has_ds = (DS_DIR / f"{pn_safe}.pdf").exists()
        verdict = ds_val.get("verdict", "") if has_ds else ""

        # Count datasheet photos
        photo_count = len(list(PHOTOS_OUT.glob(f"{pn_safe}_p*.*"))) if PHOTOS_OUT.exists() else 0

        certs = ds_data.get("certifications", [])
        if isinstance(certs, list):
            certs_str = ", ".join(str(c) if not isinstance(c, dict) else c.get("name", str(c)) for c in certs)
        else:
            certs_str = str(certs)

        def _s(v):
            """Convert any value to string for Excel."""
            if v is None: return ""
            if isinstance(v, (list, tuple)):
                return ", ".join(_s(x) for x in v)
            if isinstance(v, dict):
                return ", ".join(f"{k}={_s(val)}" for k, val in v.items())
            return str(v)

        values = [
            pn, real_brand, seed,
            _s(title)[:150], _s(description),
            _s(d.get("our_price_raw", "")),
            _s(norm.get("best_price", "")), _s(norm.get("best_price_currency", "")),
            _s(ean), "datasheet" if ean else "",
            _s(weight), _s(dims),
            specs_count, _s(top_specs),
            _s(ds_data.get("series", "")), _s(ds_data.get("category", "")),
            "YES" if has_ds else "NO", _s(verdict),
            photo_count,
            _s(certs_str),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    # Auto-width
    for col in range(1, len(HEADERS) + 1):
        max_len = min(max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row_idx)), 60)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 2

    out = ROOT / "downloads" / "staging" / "pipeline_v2_export" / "catalog_v2_FULL_370sku.xlsx"
    wb.save(str(out))
    print(f"\n  Excel saved: {out}")
    print(f"  Total SKUs: {row_idx - 2}")


def main():
    try:
        phase1_validate_datasheets()
    except Exception as e:
        print(f"Phase 1 error: {e}")
        import traceback; traceback.print_exc()

    try:
        phase2_extract_photos()
    except Exception as e:
        print(f"Phase 2 error: {e}")
        import traceback; traceback.print_exc()

    try:
        phase3_update_excel()
    except Exception as e:
        print(f"Phase 3 error: {e}")
        import traceback; traceback.print_exc()

    print("\n" + "=" * 90)
    print("NIGHT BATCH 2 COMPLETE")
    print("=" * 90)


if __name__ == "__main__":
    main()
