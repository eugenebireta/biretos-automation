"""Export 229 confirmed canonical products to Excel for owner review."""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
EV_DIR = ROOT / "downloads" / "evidence"
CANONICAL_FILE = ROOT / "downloads" / "staging" / "pipeline_v2_output" / "canonical_products.json"
OUT_FILE = ROOT / "downloads" / "staging" / "pipeline_v2_export" / "REVIEW_229_confirmed.xlsx"


def _s(v):
    if v is None: return ""
    if isinstance(v, (list, tuple)):
        return ", ".join(_s(x) for x in v[:5])
    if isinstance(v, dict):
        return ", ".join(f"{k}={_s(val)}" for k, val in list(v.items())[:5])
    return str(v)


def main():
    canonical = json.loads(CANONICAL_FILE.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "229 Confirmed Products"

    HEADERS = [
        "PN", "Brand", "Series",
        "Title (RU)",
        "Best Price", "Currency", "Price Source",
        "EAN",
        "Photo URL", "Photo Source",
        "Description (short)",
        "Specs Count", "Top Specs",
        "Weight (g)", "Dimensions (mm)", "Material",
        "Datasheet PDF",
        "Photos from datasheet",
        "Trusted Sources",
        "InSales Status", "Ozon Status", "WB Status",
    ]

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    row_idx = 2
    for p in canonical:
        identity = p.get("identity", {})
        pn = identity.get("pn", "")
        brand = identity.get("brand", "")

        # Get evidence for from_datasheet block
        pn.replace("/", "_").replace(" ", "_")
        ev_file = EV_DIR / f"evidence_{pn}.json"
        ds_block = {}
        if ev_file.exists():
            d = json.loads(ev_file.read_text(encoding="utf-8"))
            ds_block = d.get("from_datasheet", {})

        # Title - prefer datasheet
        title = ds_block.get("title") or p.get("title_ru", "")
        # Description
        desc = (p.get("best_description_ru") or "")[:300]
        # Specs
        specs = p.get("specs", {}) or ds_block.get("specs", {})
        if not isinstance(specs, dict):
            specs = {}
        # Weight/dimensions
        weight = ds_block.get("weight_g", "")
        dims = ds_block.get("dimensions_mm", "")
        material = (specs.get("Material") or specs.get("material") or
                    specs.get("MATERIAL") or "")

        # Photos
        photos_count = len(ds_block.get("product_photos", []))

        # Trusted sources
        trusted = p.get("trusted_sources", [])
        trusted_str = ", ".join(f"{s.get('domain','')}({len(s.get('has',[]))})"
                                 for s in trusted[:5])

        readiness = p.get("readiness", {})

        values = [
            pn, brand, _s(identity.get("series", "") or ds_block.get("series", "")),
            _s(title)[:150],
            _s(p.get("best_price", "")), _s(p.get("best_price_currency", "")),
            _s(p.get("best_price_source", "")),
            _s(ds_block.get("ean", "")),
            _s(p.get("best_photo_url", "")), _s(p.get("best_photo_tier", "")),
            _s(desc),
            len(specs), _s(specs)[:200],
            _s(weight), _s(dims), _s(material),
            _s(ds_block.get("datasheet_pdf", "")),
            photos_count,
            _s(trusted_str),
            _s(readiness.get("insales", "")), _s(readiness.get("ozon", "")), _s(readiness.get("wb", "")),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    # Auto-width
    for col in range(1, len(HEADERS) + 1):
        max_len = min(max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row_idx)), 60)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 2

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_FILE))
    print(f"Excel saved: {OUT_FILE}")
    print(f"Total rows: {row_idx - 2}")


if __name__ == "__main__":
    main()
