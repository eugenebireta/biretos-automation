from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from scripts.lot_scoring.pipeline.helpers import to_float, to_str

FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
FILL_ORANGE = PatternFill(fill_type="solid", start_color="FCD5B4", end_color="FCD5B4")
FILL_RED = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

HEADER_FONT = Font(bold=True)

RANKING_COLUMNS: list[tuple[str, str]] = [
    ("rank", "rank"),
    ("lot_id", "lot_id"),
    ("Tier", "Tier"),
    ("Zone_v4", "Zone_v4"),
    ("FAS_v4", "FAS_v4"),
    ("AQ_v4", "AQ_v4"),
    ("score_10", "score_10"),
    ("abs_score_100", "abs_score_100"),
    ("data_trust", "data_trust"),
    ("data_trust_pct", "data_trust_pct"),
    ("total_effective_usd", "total_effective_usd"),
    ("hhi", "hhi"),
    ("unknown_exposure", "unknown_exposure"),
    ("confidence_score", "confidence_score"),
    ("lot_avg_pn_liquidity", "lot_avg_pn_liquidity"),
    ("needs_manual_review", "needs_manual_review"),
    ("short_reason", "short_reason"),
    ("red_flags", "red_flags"),
    ("brand_detected", "brand_detected"),
    ("brand_cps", "brand_cps"),
    ("brand_flags", "brand_flags"),
    ("brand_verticals_json", "brand_verticals_json"),
]

DETAIL_COLUMNS: list[tuple[str, str]] = [
    ("rank", "rank"),
    ("lot_id", "lot_id"),
    ("Tier", "Tier"),
    ("Zone_v4", "Zone_v4"),
    ("Investor_Summary", "Investor_Summary"),
    ("Top_Loss_Reasons", "Top_Loss_Reasons"),
    ("S1", "S1"),
    ("S2", "S2"),
    ("S3", "S3"),
    ("S4", "S4"),
    ("base_score", "base_score"),
    ("final_score", "final_score"),
    ("M_hustle", "M_hustle"),
    ("M_concentration", "M_concentration"),
    ("M_tail", "M_tail"),
    ("M_Lifecycle", "M_Lifecycle"),
    ("M_Unknown", "M_Unknown"),
    ("KS_liquidity", "KS_liquidity"),
    ("KS_obsolescence", "KS_obsolescence"),
    ("data_trust", "data_trust"),
    ("data_trust_pct", "data_trust_pct"),
    ("core_ratio", "core_ratio"),
    ("core_count", "core_count"),
    ("tail_count", "tail_count"),
    ("unknown_exposure", "unknown_exposure"),
    ("unknown_exposure_usd", "unknown_exposure_usd"),
    ("needs_manual_review", "needs_manual_review"),
    ("governance_flags", "governance_flags"),
    ("flags", "flags"),
    ("brand_detected", "brand_detected"),
    ("brand_cps", "brand_cps"),
    ("brand_flags", "brand_flags"),
    ("brand_verticals_json", "brand_verticals_json"),
]


def _row_fill(row: dict[str, Any]) -> PatternFill | None:
    zone = to_str(row.get("Zone_v4"), "").upper()
    tier = to_str(row.get("Tier"), "").upper()
    unknown_exposure = to_float(row.get("unknown_exposure"), 0.0)
    needs_manual_review = bool(row.get("needs_manual_review"))
    if needs_manual_review:
        return FILL_RED
    if zone == "WEAK" or tier == "WEAK":
        return FILL_RED
    if tier == "RISKY" or unknown_exposure >= 0.30:
        return FILL_ORANGE
    if zone == "REVIEW" or tier == "CANDIDATE":
        return FILL_YELLOW
    if zone == "PASS" and tier == "STRONG":
        return FILL_GREEN
    return None


def _cell_fill(column_key: str, row: dict[str, Any]) -> PatternFill | None:
    if column_key == "unknown_exposure":
        unknown_exposure = to_float(row.get("unknown_exposure"), 0.0)
        if unknown_exposure >= 0.40:
            return FILL_RED
        if unknown_exposure >= 0.30:
            return FILL_ORANGE
    if column_key == "hhi":
        hhi = to_float(row.get("hhi"), 0.0)
        if hhi >= 0.60:
            return FILL_RED
    if column_key == "data_trust":
        trust = to_str(row.get("data_trust"), "").upper()
        if trust == "LOW":
            return FILL_ORANGE
    if column_key == "needs_manual_review" and bool(row.get("needs_manual_review")):
        return FILL_RED
    if column_key == "brand_cps":
        brand_cps = to_float(row.get("brand_cps"), 0.0)
        if brand_cps < 0.0:
            return FILL_RED
        if brand_cps < 0.35:
            return FILL_ORANGE
    return None


def _format_cell(cell: Any, key: str) -> None:
    if key in {"score_10", "FAS_v4", "AQ_v4", "abs_score_100"}:
        cell.number_format = "0.00"
        return
    if key in {
        "unknown_exposure",
        "data_trust_pct",
        "hhi",
        "confidence_score",
        "lot_avg_pn_liquidity",
        "core_ratio",
        "S1",
        "S2",
        "S3",
        "S4",
        "M_hustle",
        "M_concentration",
        "M_tail",
        "M_Lifecycle",
        "M_Unknown",
        "KS_liquidity",
        "KS_obsolescence",
        "brand_cps",
    }:
        cell.number_format = "0.000"
        return
    if key in {"total_effective_usd", "unknown_exposure_usd"}:
        cell.number_format = "0.00"


def _apply_auto_width(ws: Any) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 48)


def _write_sheet(ws: Any, rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> None:
    for col_idx, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(rows, start=2):
        row_fill = _row_fill(row)
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _format_cell(cell, key)
            if row_fill is not None:
                cell.fill = row_fill
            forced_fill = _cell_fill(key, row)
            if forced_fill is not None:
                cell.fill = forced_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _apply_auto_width(ws)


def write_investor_report(full_rows: list[dict[str, Any]], output_path: Path) -> Path:
    ranked_rows = sorted(
        [dict(row) for row in full_rows],
        key=lambda row: int(to_float(row.get("rank"), 0.0)),
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ranking_sheet = workbook.create_sheet(title="RANKING")
    _write_sheet(ranking_sheet, ranked_rows, RANKING_COLUMNS)

    detail_sheet = workbook.create_sheet(title="DETAIL")
    _write_sheet(detail_sheet, ranked_rows, DETAIL_COLUMNS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
    return output_path
