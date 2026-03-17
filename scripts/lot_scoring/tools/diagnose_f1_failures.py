from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _normalize_lot_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.lower().startswith("lot_"):
        text = text[4:]
    if text.lower().endswith("_artifact"):
        text = text[:-9]
    if text.replace(".", "", 1).isdigit():
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if parsed != parsed:
        return None
    return parsed


def _fmt_num(value: Any, precision: int = 6) -> str:
    parsed = _to_float(value)
    if parsed is None:
        return "N/A"
    if parsed.is_integer():
        return str(int(parsed))
    return f"{parsed:.{precision}f}"


def _fmt_money(value: Any) -> str:
    parsed = _to_float(value)
    if parsed is None:
        return "N/A"
    return f"{parsed:,.2f}".replace(",", " ")


def _escape_md(value: Any) -> str:
    if value is None:
        return "N/A"
    text = str(value).replace("\n", " ").strip()
    if not text:
        return "N/A"
    return text.replace("|", "\\|")


def _table(headers: list[str], rows: list[list[Any]]) -> str:
    if not rows:
        return "N/A"
    out = [
        "| " + " | ".join(_escape_md(h) for h in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(_escape_md(cell) for cell in row) + " |")
    return "\n".join(out)


def _load_summary_rows(summary_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(summary_path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_normalize_header(col) for col in header_row]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)
    return headers, rows


def _load_artifact(downloads_dir: Path, lot_id: str) -> dict[str, Any]:
    candidate = downloads_dir / f"lot_{lot_id}_artifact.json"
    if not candidate.exists():
        return {}
    try:
        with candidate.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
            return payload if isinstance(payload, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _safe_get(
    summary_row: dict[str, Any],
    summary_key: str,
    artifact: dict[str, Any],
    metric_key: str | None = None,
    status_key: str | None = None,
) -> Any:
    if summary_key in summary_row and summary_row.get(summary_key) is not None:
        return summary_row.get(summary_key)
    if status_key:
        statuses = artifact.get("statuses")
        if isinstance(statuses, dict):
            status_value = statuses.get(status_key)
            if status_value is not None:
                return status_value
    metrics = artifact.get("metrics")
    if metric_key and isinstance(metrics, dict):
        metric_value = metrics.get(metric_key)
        if metric_value is not None:
            return metric_value
    return None


def _f1_fail_reasons(share_ab: float | None, share_e: float | None, largest: float | None) -> list[str]:
    reasons: list[str] = []
    if share_ab is not None and share_ab < 0.60 - 1e-12:
        reasons.append("FAIL_AB")
    if share_e is not None and share_e > 0.20 + 1e-12:
        reasons.append("FAIL_E")
    if largest is not None and largest > 0.45 + 1e-12:
        reasons.append("FAIL_LARGEST")
    return reasons


def _near_miss(
    f1_status: str,
    share_ab: float | None,
    share_e: float | None,
    largest: float | None,
) -> bool:
    if f1_status.upper() != "STOP":
        return False
    checks: list[bool] = []
    if share_ab is not None and share_ab < 0.60:
        checks.append(share_ab >= 0.57)
    if share_e is not None and share_e > 0.20:
        checks.append(share_e <= 0.23)
    if largest is not None and largest > 0.45:
        checks.append(largest <= 0.48)
    return any(checks)


def _shortfall_sum(margin_ab: float | None, margin_e: float | None, margin_largest: float | None) -> float:
    def neg_part(value: float | None) -> float:
        return max(0.0, -(value or 0.0))

    return neg_part(margin_ab) + neg_part(margin_e) + neg_part(margin_largest)


def build_reports(summary_path: Path, downloads_dir: Path) -> tuple[Path, Path]:
    _, rows = _load_summary_rows(summary_path)
    diagnostics: list[dict[str, Any]] = []
    for row in rows:
        lot_id = _normalize_lot_id(row.get("lot_id"))
        if not lot_id:
            continue
        artifact = _load_artifact(downloads_dir, lot_id)

        f0_status = str(_safe_get(row, "f0_status", artifact, status_key="f0") or "N/A")
        f1_status = str(_safe_get(row, "f1_status", artifact, status_key="f1") or "N/A")
        f2_status = str(_safe_get(row, "f2_status", artifact, status_key="f2") or "N/A")

        total_value_rub = _to_float(_safe_get(row, "total_value_rub", artifact, metric_key="total_value_rub"))
        share_a = _to_float(_safe_get(row, "share_a", artifact, metric_key="share_a"))
        share_b = _to_float(_safe_get(row, "share_b", artifact, metric_key="share_b"))
        share_e = _to_float(_safe_get(row, "share_e", artifact, metric_key="share_e"))
        largest = _to_float(_safe_get(row, "largest_sku_share", artifact, metric_key="largest_sku_share"))
        ticket_pct = _to_float(_safe_get(row, "ticket_pct", artifact, metric_key="ticket_pct"))
        mass_anchor_status = str(
            _safe_get(row, "mass_anchor_status", artifact, metric_key="mass_anchor_status") or "N/A"
        )

        share_ab = None if (share_a is None or share_b is None) else (share_a + share_b)
        margin_e = None if share_e is None else (0.20 - share_e)
        margin_ab = None if share_ab is None else (share_ab - 0.60)
        margin_largest = None if largest is None else (0.45 - largest)
        reasons = _f1_fail_reasons(share_ab, share_e, largest)
        near_miss = _near_miss(f1_status, share_ab, share_e, largest)
        shortfall = _shortfall_sum(margin_ab, margin_e, margin_largest)

        diagnostics.append(
            {
                "lot_id": lot_id,
                "f0_status": f0_status,
                "f1_status": f1_status,
                "f2_status": f2_status,
                "total_value_rub": total_value_rub,
                "share_ab": share_ab,
                "share_e": share_e,
                "largest": largest,
                "ticket_pct": ticket_pct,
                "mass_anchor_status": mass_anchor_status,
                "margin_ab": margin_ab,
                "margin_e": margin_e,
                "margin_largest": margin_largest,
                "reasons": reasons,
                "near_miss": near_miss,
                "shortfall_sum": shortfall,
            }
        )

    diagnostics.sort(
        key=lambda item: (_to_float(item.get("total_value_rub")) or float("-inf")),
        reverse=True,
    )

    diagnostic_path = downloads_dir / "F1_DIAGNOSTIC.md"
    shortlist_path = downloads_dir / "NEAR_MISS_SHORTLIST.md"

    diag_rows: list[list[Any]] = []
    for item in diagnostics:
        diag_rows.append(
            [
                item["lot_id"],
                item["f0_status"],
                item["f1_status"],
                item["f2_status"],
                _fmt_money(item["total_value_rub"]),
                _fmt_num(item["share_ab"], 6),
                _fmt_num(item["share_e"], 6),
                _fmt_num(item["largest"], 6),
                ", ".join(item["reasons"]) if item["reasons"] else "PASS",
                str(bool(item["near_miss"])).lower(),
            ]
        )

    diagnostic_md = [
        "# F1_DIAGNOSTIC",
        "",
        "All lots sorted by total_value_rub (desc).",
        "",
        _table(
            [
                "lot_id",
                "f0",
                "f1",
                "f2",
                "total_value_rub",
                "share_ab",
                "share_e",
                "largest",
                "reasons",
                "near_miss",
            ],
            diag_rows,
        ),
        "",
    ]
    diagnostic_path.write_text("\n".join(diagnostic_md), encoding="utf-8")

    near_miss_items = [item for item in diagnostics if item["near_miss"]]
    near_miss_items.sort(key=lambda item: (item["shortfall_sum"], -(item["total_value_rub"] or 0.0)))

    shortlist_md: list[str] = []
    shortlist_md.append("# NEAR_MISS_SHORTLIST")
    shortlist_md.append("")
    shortlist_md.append("Near-miss lots (F1=STOP and at least one failed condition within 3 p.p. from threshold).")
    shortlist_md.append("")
    if not near_miss_items:
        shortlist_md.append("No near-miss lots found.")
        shortlist_md.append("")
    else:
        for item in near_miss_items:
            lot_id = item["lot_id"]
            shortlist_md.append(f"## LOT {lot_id}")
            shortlist_md.append("")
            shortlist_md.append(
                f"- f0/f1/f2: `{item['f0_status']}` / `{item['f1_status']}` / `{item['f2_status']}`; "
                f"total_value_rub: `{_fmt_money(item['total_value_rub'])}`"
            )
            shortlist_md.append(f"- shortfall_sum: `{_fmt_num(item['shortfall_sum'], 6)}`")
            shortlist_md.append("- failed_conditions:")

            reasons = item["reasons"]
            if "FAIL_AB" in reasons:
                pp = max(0.0, (0.60 - (item["share_ab"] or 0.0)) * 100.0)
                shortlist_md.append(
                    f"  - FAIL_AB: share_ab=`{_fmt_num(item['share_ab'], 6)}`; "
                    f"below threshold by `{_fmt_num(pp, 2)}` p.p."
                )
            if "FAIL_E" in reasons:
                pp = max(0.0, ((item["share_e"] or 0.0) - 0.20) * 100.0)
                shortlist_md.append(
                    f"  - FAIL_E: share_e=`{_fmt_num(item['share_e'], 6)}`; "
                    f"above threshold by `{_fmt_num(pp, 2)}` p.p."
                )
                shortlist_md.append(
                    f"  - hint: inspect `downloads/lots/lot_{lot_id}_artifact.json` -> "
                    f"`top80_value_slice` and `review_lines` for unknown_category/unknown_brand drivers."
                )
            if "FAIL_LARGEST" in reasons:
                pp = max(0.0, ((item["largest"] or 0.0) - 0.45) * 100.0)
                shortlist_md.append(
                    f"  - FAIL_LARGEST: largest_sku_share=`{_fmt_num(item['largest'], 6)}`; "
                    f"above threshold by `{_fmt_num(pp, 2)}` p.p."
                )
            shortlist_md.append("")

    shortlist_path.write_text("\n".join(shortlist_md), encoding="utf-8")
    return diagnostic_path, shortlist_path


def _default_paths(repo_root: Path) -> tuple[Path, Path]:
    downloads_dir = repo_root / "downloads" / "lots"
    summary_path = downloads_dir / "summary_filters_report.xlsx"
    return summary_path, downloads_dir


def main() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    default_summary, default_downloads = _default_paths(repo_root)

    parser = argparse.ArgumentParser(description="Diagnose F1 failures and near-miss lots from Compute Engine v3 outputs.")
    parser.add_argument("--summary", default=str(default_summary), help="Path to summary_filters_report.xlsx")
    parser.add_argument("--downloads-dir", default=str(default_downloads), help="Path to downloads/lots directory")
    args = parser.parse_args()

    summary_path = Path(args.summary)
    downloads_dir = Path(args.downloads_dir)
    if not summary_path.exists():
        raise FileNotFoundError(f"Summary file not found: {summary_path}")
    if not downloads_dir.exists():
        raise FileNotFoundError(f"Downloads directory not found: {downloads_dir}")

    diagnostic_path, shortlist_path = build_reports(summary_path, downloads_dir)
    print(f"F1 diagnostic written to: {diagnostic_path}")
    print(f"Near-miss shortlist written to: {shortlist_path}")


if __name__ == "__main__":
    main()

# Run:
# python -m scripts.lot_scoring.tools.diagnose_f1_failures
