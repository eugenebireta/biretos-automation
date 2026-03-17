from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

QTY_TEXT_PATTERN = re.compile(r"кол-во\s*([0-9]+(?:[.,][0-9]+)?)", re.IGNORECASE)


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _normalize_lot_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        return ""
    if text.lower().startswith("lot_"):
        text = text[4:]
    if text.lower().endswith("_artifact"):
        text = text[:-9]
    if re.fullmatch(r"\d+(\.0+)?", text):
        return str(int(float(text)))
    if text.isdigit():
        return str(int(text))
    match = re.search(r"\d+", text)
    if match:
        return str(int(match.group(0)))
    return text


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _fmt_num(value: Any, precision: int = 4) -> str:
    if value is None:
        return "N/A"
    parsed = _to_float(value)
    if parsed is None:
        text = str(value).strip()
        return text if text else "N/A"
    if parsed.is_integer():
        return str(int(parsed))
    return f"{parsed:.{precision}f}"


def _fmt_money(value: Any) -> str:
    parsed = _to_float(value)
    if parsed is None:
        return "N/A"
    return f"{parsed:,.2f}".replace(",", " ")


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "да"}


def _escape_md(value: Any) -> str:
    if value is None:
        return "N/A"
    text = str(value).replace("\n", " ").strip()
    if not text:
        return "N/A"
    return text.replace("|", "\\|")


def _extract_qty(row: dict[str, Any]) -> float | None:
    # Strict policy: raw_qty -> qty -> total_qty, no SKU parsing.
    for key in ("raw_qty", "qty", "total_qty"):
        parsed = _to_float(row.get(key))
        if parsed is not None:
            return parsed
    return None


def _extract_qty_from_text(raw_text: Any) -> float | None:
    if raw_text is None:
        return None
    text = str(raw_text)
    match = QTY_TEXT_PATTERN.search(text)
    if not match:
        return None
    raw = match.group(1).replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def _is_suspicious_qty_mismatch(row: dict[str, Any]) -> bool:
    qty = _extract_qty(row)
    if qty is None:
        return False
    text_qty = _extract_qty_from_text(row.get("raw_text"))
    if text_qty is None:
        return False
    return abs(qty - text_qty) > 1e-9


def _load_summary_rows(summary_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(summary_path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_normalize_header(col) for col in header_row]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)
    return headers, rows


def _select_candidate_lots(headers: list[str], rows: list[dict[str, Any]], top_n: int = 5) -> list[str]:
    has_candidate_col = "filter4_candidate" in headers
    selected_rows: list[dict[str, Any]]
    if has_candidate_col:
        selected_rows = [row for row in rows if _as_bool(row.get("filter4_candidate"))]
    else:
        eligible = [
            row
            for row in rows
            if str(row.get("f1_status", "")).upper() == "PASS" and str(row.get("f2_status", "")).upper() != "STOP"
        ]
        selected_rows = sorted(
            eligible,
            key=lambda row: (_to_float(row.get("score_100")) or float("-inf")),
            reverse=True,
        )[:top_n]
    lot_ids: list[str] = []
    for row in selected_rows:
        lot_id = _normalize_lot_id(row.get("lot_id"))
        if lot_id and lot_id not in lot_ids:
            lot_ids.append(lot_id)
    return lot_ids


def _load_artifact(downloads_dir: Path, lot_id: str) -> dict[str, Any]:
    direct = downloads_dir / f"lot_{lot_id}_artifact.json"
    if direct.exists():
        with direct.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    for candidate in sorted(downloads_dir.glob("lot_*_artifact.json")):
        if _normalize_lot_id(candidate.stem) == lot_id or f"lot_{lot_id}_artifact" == candidate.stem:
            with candidate.open("r", encoding="utf-8") as handle:
                return json.load(handle)
    return {}


def _table(headers: list[str], rows: list[list[Any]]) -> str:
    if not rows:
        return "N/A"
    out = []
    out.append("| " + " | ".join(_escape_md(h) for h in headers) + " |")
    out.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        out.append("| " + " | ".join(_escape_md(cell) for cell in row) + " |")
    return "\n".join(out)


def _slice_table(slice_rows: list[dict[str, Any]], limit: int) -> str:
    if not slice_rows:
        return "N/A"
    rows: list[list[Any]] = []
    for idx, item in enumerate(slice_rows[:limit], start=1):
        qty = _extract_qty(item)
        suspicious = _is_suspicious_qty_mismatch(item)
        rows.append(
            [
                idx,
                item.get("sku", "N/A"),
                item.get("raw_text", "N/A"),
                _fmt_num(qty),
                _fmt_money(item.get("line_value_rub")),
                item.get("category", "N/A"),
                item.get("brand", "N/A"),
                "true" if suspicious else "false",
            ]
        )
    table = _table(
        ["#", "sku", "raw_text", "qty", "line_value_rub", "category", "brand", "suspicious_qty_mismatch"],
        rows,
    )
    if len(slice_rows) > limit:
        table += f"\n\n... truncated ... ({len(slice_rows) - limit} more rows)"
    return table


def _review_lines_table(review_lines: list[dict[str, Any]], limit: int) -> str:
    if not review_lines:
        return "N/A"
    rows: list[list[Any]] = []
    for idx, item in enumerate(review_lines[:limit], start=1):
        reasons = item.get("reasons")
        if isinstance(reasons, list):
            reasons_text = ", ".join(str(x) for x in reasons)
        else:
            reasons_text = "N/A"
        qty = _extract_qty(item)
        suspicious = _is_suspicious_qty_mismatch(item)
        rows.append(
            [
                idx,
                item.get("sku", "N/A"),
                item.get("raw_text", "N/A"),
                _fmt_num(qty),
                _fmt_money(item.get("line_value_rub")),
                reasons_text,
                "true" if suspicious else "false",
            ]
        )
    table = _table(["#", "sku", "raw_text", "qty", "line_value_rub", "reasons", "suspicious_qty_mismatch"], rows)
    if len(review_lines) > limit:
        table += f"\n\n... truncated ... ({len(review_lines) - limit} more rows)"
    return table


def _value_from_sources(
    summary_row: dict[str, Any] | None,
    metrics: dict[str, Any],
    key: str,
) -> Any:
    if summary_row and key in summary_row and summary_row.get(key) is not None:
        return summary_row.get(key)
    return metrics.get(key)


def _build_filter4_checklist(checklist: dict[str, Any]) -> str:
    fatal = checklist.get("fatal_items") if isinstance(checklist, dict) else None
    market = checklist.get("market_rule_top3") if isinstance(checklist, dict) else None
    physical = checklist.get("physical_degradation_rule") if isinstance(checklist, dict) else None
    adaptive = checklist.get("adaptive_depth") if isinstance(checklist, dict) else None

    lines: list[str] = []
    lines.append("### Filter 4 Checklist")
    lines.append("")
    lines.append("#### 4A. Critical constraints")
    if isinstance(fatal, list) and fatal:
        for item in fatal:
            rule = item.get("rule", "N/A") if isinstance(item, dict) else "N/A"
            confirmed = item.get("if_confirmed", "STOP") if isinstance(item, dict) else "STOP"
            unknown = item.get("if_unknown", "REVIEW") if isinstance(item, dict) else "REVIEW"
            lines.append(f"- [ ] {rule}: confirmed -> {confirmed}, unknown -> {unknown}")
    else:
        lines.extend(
            [
                "- [ ] vendor_lock_or_firmware_lock: confirmed -> STOP, unknown -> REVIEW",
                "- [ ] legal_non_resalable: confirmed -> STOP, unknown -> REVIEW",
                "- [ ] critical_revision_incompatibility_or_project_lock: confirmed -> STOP, unknown -> REVIEW",
            ]
        )

    lines.append("")
    lines.append("#### 4B. Market rule")
    if isinstance(market, list) and market:
        for rule in market:
            lines.append(f"- [ ] {rule}")
    else:
        lines.extend(
            [
                "- [ ] dead_market_top3_count>=2 => STOP",
                "- [ ] dead_market_top3_count==1 => REVIEW",
                "- [ ] else => PASS",
            ]
        )

    lines.append("")
    lines.append("#### 4C. Degradation / shelf-life")
    if isinstance(physical, list) and physical:
        for rule in physical:
            lines.append(f"- [ ] {rule}")
    else:
        lines.extend(
            [
                "- [ ] high_risk_share>20% => REVIEW",
                "- [ ] (high_risk_share+medium_risk_share)>30% => REVIEW/SKIP",
            ]
        )
    if isinstance(adaptive, dict):
        base_x = adaptive.get("base_x", "N/A")
        lines.append(f"- [ ] adaptive_depth_base_x={base_x}")
    lines.append("- [ ] shelf-risk signals assessed (if present: raise depth to X=80)")
    return "\n".join(lines)


def build_pack(
    summary_path: Path,
    downloads_dir: Path,
    output_path: Path,
    top_n_fallback: int = 5,
) -> list[str]:
    headers, summary_rows = _load_summary_rows(summary_path)
    summary_by_lot = {_normalize_lot_id(row.get("lot_id")): row for row in summary_rows}
    selected_lots = _select_candidate_lots(headers, summary_rows, top_n=top_n_fallback)

    lines: list[str] = []
    lines.append("# FILTER4_PACK")
    lines.append("")
    lines.append("Generated from Compute Engine v3 artifacts.")
    lines.append("")

    if not selected_lots:
        lines.append("## Selected Lots")
        lines.append("")
        lines.append("No Filter 4 candidates were selected by current rules.")
        output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return []

    lines.append("## Candidate Summary")
    lines.append("")
    summary_table_rows: list[list[Any]] = []

    enriched: list[tuple[str, dict[str, Any], dict[str, Any] | None]] = []
    for lot_id in selected_lots:
        artifact = _load_artifact(downloads_dir, lot_id)
        summary_row = summary_by_lot.get(lot_id)
        metrics = artifact.get("metrics", {}) if isinstance(artifact, dict) else {}
        score = _value_from_sources(summary_row, metrics, "score_100")
        summary_table_rows.append(
            [
                lot_id,
                _fmt_num(score, precision=0),
                _value_from_sources(summary_row, metrics, "class") or "N/A",
                _fmt_money(_value_from_sources(summary_row, metrics, "total_value_rub")),
                _fmt_num(_value_from_sources(summary_row, metrics, "share_e"), precision=6),
                _fmt_num(_value_from_sources(summary_row, metrics, "largest_sku_share"), precision=6),
                _value_from_sources(summary_row, metrics, "mass_anchor_status") or "N/A",
            ]
        )
        enriched.append((lot_id, artifact, summary_row))

    lines.append(
        _table(
            ["lot_id", "score_100", "class", "total_value_rub", "share_e", "largest_sku_share", "mass_anchor_status"],
            summary_table_rows,
        )
    )
    lines.append("")

    for lot_id, artifact, summary_row in enriched:
        metrics = artifact.get("metrics", {}) if isinstance(artifact, dict) else {}
        statuses = artifact.get("statuses", {}) if isinstance(artifact, dict) else {}
        top3 = artifact.get("top3_by_value", []) if isinstance(artifact, dict) else []
        top20 = artifact.get("top20_value_slice", []) if isinstance(artifact, dict) else []
        top40 = artifact.get("top40_value_slice", []) if isinstance(artifact, dict) else []
        top80 = artifact.get("top80_value_slice", []) if isinstance(artifact, dict) else []
        review_lines = artifact.get("review_lines", []) if isinstance(artifact, dict) else []
        checklist = artifact.get("filter4_checklist", {}) if isinstance(artifact, dict) else {}

        lines.append(f"## LOT {lot_id}")
        lines.append("")
        lines.append("### Key Metrics")
        lines.append("")
        lines.append(
            f"- score_100: `{_fmt_num(_value_from_sources(summary_row, metrics, 'score_100'), precision=0)}`; "
            f"class: `{_value_from_sources(summary_row, metrics, 'class') or 'N/A'}`"
        )
        lines.append(
            f"- share_a/share_b/share_e: `{_fmt_num(_value_from_sources(summary_row, metrics, 'share_a'), 6)}` / "
            f"`{_fmt_num(_value_from_sources(summary_row, metrics, 'share_b'), 6)}` / "
            f"`{_fmt_num(_value_from_sources(summary_row, metrics, 'share_e'), 6)}`"
        )
        lines.append(
            f"- largest_sku_share: `{_fmt_num(_value_from_sources(summary_row, metrics, 'largest_sku_share'), 6)}`; "
            f"core70_value_rub: `{_fmt_money(_value_from_sources(summary_row, metrics, 'core70_value_rub'))}`"
        )
        lines.append(
            f"- entry_price_rub: `{_fmt_money(_value_from_sources(summary_row, metrics, 'entry_price_rub'))}`; "
            f"mass_anchor_status: `{_value_from_sources(summary_row, metrics, 'mass_anchor_status') or 'N/A'}`"
        )
        lines.append(
            f"- ticket_pct: `{_fmt_num(_value_from_sources(summary_row, metrics, 'ticket_pct'), 6)}`; "
            f"used_share: `{_fmt_num(_value_from_sources(summary_row, metrics, 'share_used'), 6)}`"
        )
        lines.append(
            f"- statuses f0/f1/f2: `{statuses.get('f0', 'N/A')}` / `{statuses.get('f1', 'N/A')}` / `{statuses.get('f2', 'N/A')}`"
        )
        lines.append("")

        lines.append("### Top-3 by Value")
        lines.append("")
        lines.append(_slice_table(top3 if isinstance(top3, list) else [], limit=10))
        lines.append("")

        lines.append("### Top20 Value Slice")
        lines.append("")
        lines.append(_slice_table(top20 if isinstance(top20, list) else [], limit=10))
        lines.append("")

        lines.append("### Top40 Value Slice")
        lines.append("")
        lines.append(_slice_table(top40 if isinstance(top40, list) else [], limit=10))
        lines.append("")

        lines.append("### Top80 Value Slice")
        lines.append("")
        lines.append(_slice_table(top80 if isinstance(top80, list) else [], limit=10))
        lines.append("")

        lines.append("### Review Lines")
        lines.append("")
        lines.append(_review_lines_table(review_lines if isinstance(review_lines, list) else [], limit=25))
        lines.append("")

        lines.append(_build_filter4_checklist(checklist if isinstance(checklist, dict) else {}))
        lines.append("")

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return selected_lots


def build_top10_lots(summary_path: Path, output_path: Path) -> None:
    _, rows = _load_summary_rows(summary_path)
    sorted_rows = sorted(
        rows,
        key=lambda row: (_to_float(row.get("score_100")) if _to_float(row.get("score_100")) is not None else float("-inf")),
        reverse=True,
    )
    top10 = sorted_rows[:10]

    lines: list[str] = []
    lines.append("# TOP10_LOTS")
    lines.append("")
    lines.append("Top-10 lots by score_100 (desc).")
    lines.append("")
    table_rows: list[list[Any]] = []
    for row in top10:
        table_rows.append(
            [
                _normalize_lot_id(row.get("lot_id")) or "N/A",
                _fmt_num(row.get("score_100"), precision=0),
                row.get("class", "N/A"),
                row.get("f0_status", "N/A"),
                row.get("f1_status", "N/A"),
                row.get("f2_status", "N/A"),
                _fmt_num(row.get("share_e"), precision=6),
                _fmt_num(row.get("largest_sku_share"), precision=6),
                row.get("mass_anchor_status", "N/A"),
                _fmt_num(row.get("ticket_pct"), precision=6),
                _fmt_money(row.get("total_value_rub")),
            ]
        )
    lines.append(
        _table(
            [
                "lot_id",
                "score_100",
                "class",
                "f0_status",
                "f1_status",
                "f2_status",
                "share_e",
                "largest_sku_share",
                "mass_anchor_status",
                "ticket_pct",
                "total_value_rub",
            ],
            table_rows,
        )
    )
    lines.append("")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def _default_paths(repo_root: Path) -> tuple[Path, Path, Path, Path]:
    downloads_dir = repo_root / "downloads" / "lots"
    summary_path = downloads_dir / "summary_filters_report.xlsx"
    filter4_output_path = downloads_dir / "FILTER4_PACK.md"
    top10_output_path = downloads_dir / "TOP10_LOTS.md"
    return summary_path, downloads_dir, filter4_output_path, top10_output_path


def main() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    default_summary, default_downloads, default_filter4_output, default_top10_output = _default_paths(repo_root)

    parser = argparse.ArgumentParser(description="Build compact Filter 4 markdown pack from existing artifacts.")
    parser.add_argument("--summary", default=str(default_summary), help="Path to summary_filters_report.xlsx")
    parser.add_argument("--downloads-dir", default=str(default_downloads), help="Path to downloads/lots directory")
    parser.add_argument("--output", default=str(default_filter4_output), help="Path to output FILTER4_PACK.md")
    parser.add_argument("--top10-output", default=str(default_top10_output), help="Path to output TOP10_LOTS.md")
    parser.add_argument("--top-n-fallback", type=int, default=5, help="Fallback top-N if filter4_candidate is absent")
    args = parser.parse_args()

    summary_path = Path(args.summary)
    downloads_dir = Path(args.downloads_dir)
    output_path = Path(args.output)
    top10_output_path = Path(args.top10_output)

    if not summary_path.exists():
        raise FileNotFoundError(f"Summary file not found: {summary_path}")
    if not downloads_dir.exists():
        raise FileNotFoundError(f"Downloads directory not found: {downloads_dir}")

    selected_lots = build_pack(
        summary_path=summary_path,
        downloads_dir=downloads_dir,
        output_path=output_path,
        top_n_fallback=max(1, int(args.top_n_fallback)),
    )
    build_top10_lots(summary_path=summary_path, output_path=top10_output_path)
    print(f"FILTER4 pack written to: {output_path}")
    print(f"TOP10 lots written to: {top10_output_path}")
    print(f"Selected lot_ids in FILTER4_PACK: {selected_lots}")
    print("Qty policy: raw_qty -> qty -> total_qty; no fallback from SKU digits.")


if __name__ == "__main__":
    main()
