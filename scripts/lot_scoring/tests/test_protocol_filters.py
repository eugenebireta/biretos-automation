from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.lot_scoring.filters.artifact import build_lot_artifact_payload, write_lot_artifact
from scripts.lot_scoring.filters.engine import run_filters
from scripts.lot_scoring.filters.f1_capital_composition import evaluate_f1
from scripts.lot_scoring.filters.f2_capital_safety import evaluate_f2
from scripts.lot_scoring.filters.f3_structural_ranking import evaluate_f3
from scripts.lot_scoring.filters.protocol_config import ProtocolConfig
from scripts.lot_scoring.filters.protocol_metrics import (
    aggregate_by_sku,
    core70,
    prepare_protocol_lines,
    shares,
    ticket_pct,
    total_value,
    used_stress_adjust,
)


def _sku(
    sku: str,
    value_rub: float,
    *,
    category: str,
    qty: float = 1.0,
    brand: str = "brandx",
    condition_raw: float = 1.0,
    base_unit_usd: float = 10.0,
) -> dict:
    return {
        "sku_code_normalized": sku,
        "CalculatedLotCost": value_rub,
        "raw_qty": qty,
        "category": category,
        "brand": brand,
        "condition_raw": condition_raw,
        "base_unit_usd": base_unit_usd,
        "effective_line_usd": max(0.0, value_rub / 100.0),
    }


def test_share_sum_base_and_adjusted_equals_one() -> None:
    cfg = ProtocolConfig()
    rows = [
        _sku("A1", 60_000, category="gas_safety", condition_raw=0.8),
        _sku("B1", 25_000, category="hvac_components", condition_raw=0.8),
        _sku("E1", 15_000, category="it_hardware"),
    ]
    lines = prepare_protocol_lines(rows, usd_rate=100.0, config=cfg)
    total = total_value(lines)
    base = shares(lines, cfg, total, value_key="line_value_rub")
    assert pytest.approx(base["share_a"] + base["share_b"] + base["share_e"], rel=1e-9) == 1.0

    used_info = used_stress_adjust(lines, cfg)
    adjusted_total = float(used_info["adjusted_total_value_rub"])
    adjusted = shares(lines, cfg, adjusted_total, value_key="adjusted_value_rub")
    assert pytest.approx(adjusted["share_a"] + adjusted["share_b"] + adjusted["share_e"], rel=1e-9) == 1.0


def test_largest_share_uses_aggregated_sku_value() -> None:
    cfg = ProtocolConfig()
    rows = [
        _sku("DUP", 100_000, category="gas_safety"),
        _sku("DUP", 100_000, category="gas_safety"),
        _sku("ONE", 100_000, category="hvac_components"),
    ]
    lines = prepare_protocol_lines(rows, usd_rate=100.0, config=cfg)
    total = total_value(lines)
    f1 = evaluate_f1(lines=lines, total_value_rub=total, match_confidence_weighted=None, config=cfg)
    assert pytest.approx(f1["largest_sku_share"], rel=1e-9) == (200_000.0 / 300_000.0)


def test_core70_is_accumulated_until_threshold() -> None:
    cfg = ProtocolConfig()
    rows = [
        _sku("A", 50_000, category="gas_safety"),
        _sku("B", 30_000, category="gas_safety"),
        _sku("C", 20_000, category="gas_safety"),
    ]
    lines = prepare_protocol_lines(rows, usd_rate=100.0, config=cfg)
    aggregated = aggregate_by_sku(lines)
    total = total_value(lines)
    core_slice, core_value = core70(aggregated, total, cfg)
    assert core_value >= 0.70 * total
    assert len(core_slice) == 2


def test_f1_binary_pass_or_stop_only() -> None:
    cfg = ProtocolConfig()
    pass_rows = [
        _sku("A", 60_000, category="gas_safety"),
        _sku("B", 20_000, category="hvac_components"),
        _sku("E", 20_000, category="unknown"),
    ]
    stop_rows = [
        _sku("A", 30_000, category="gas_safety"),
        _sku("E1", 35_000, category="unknown"),
        _sku("E2", 35_000, category="it_hardware"),
    ]
    pass_lines = prepare_protocol_lines(pass_rows, usd_rate=100.0, config=cfg)
    stop_lines = prepare_protocol_lines(stop_rows, usd_rate=100.0, config=cfg)
    f1_pass = evaluate_f1(lines=pass_lines, total_value_rub=total_value(pass_lines), match_confidence_weighted=None, config=cfg)
    f1_stop = evaluate_f1(lines=stop_lines, total_value_rub=total_value(stop_lines), match_confidence_weighted=None, config=cfg)
    assert f1_pass["status"] in {"PASS", "STOP"}
    assert f1_stop["status"] in {"PASS", "STOP"}
    assert f1_stop["status"] == "STOP"


def test_ticket_pct_counts_unique_aggregated_sku() -> None:
    cfg = ProtocolConfig()
    rows = [
        _sku("DUP", 60_000, category="gas_safety"),
        _sku("DUP", 60_000, category="gas_safety"),
        _sku("LOW", 50_000, category="gas_safety"),
    ]
    lines = prepare_protocol_lines(rows, usd_rate=100.0, config=cfg)
    aggregated = aggregate_by_sku(lines)
    assert pytest.approx(ticket_pct(aggregated, 100_000.0), rel=1e-9) == 0.5


def test_entry_price_absent_sets_f2_review_and_f3_zero() -> None:
    cfg = ProtocolConfig()
    rows = [
        _sku("A", 500_000, category="gas_safety"),
        _sku("B", 500_000, category="hvac_components"),
    ]
    lines = prepare_protocol_lines(rows, usd_rate=100.0, config=cfg)
    total = total_value(lines)
    aggregated = aggregate_by_sku(lines)
    f1 = evaluate_f1(lines=lines, total_value_rub=total, match_confidence_weighted=None, config=cfg)
    f2 = evaluate_f2(aggregated=aggregated, total_value_rub=total, entry_price_rub=None, config=cfg)
    f3 = evaluate_f3(
        f1_status=f1["status"],
        f2_status=f2["status"],
        share_a=f1["share_a"],
        share_b=f1["share_b"],
        share_e=f1["share_e"],
        total_value_rub=total,
        entry_price_rub=None,
        largest_sku_share=f2["largest_sku_share"],
        match_confidence_weighted=None,
        core70_value_rub=f2["core70_value_rub"],
        share_used=f1["share_used"],
        used_signal_available=True,
        ticket_pct=ticket_pct(aggregated, cfg.f3_ticket_threshold_rub),
        core_sku_count=1,
        brand_count_core=1,
    )
    assert f2["status"] == "REVIEW"
    assert "NEEDS_ENTRYPRICE" in f2["review_reasons"]
    assert f3["score_100"] == 0
    assert f3["class"] == "Weak"


def test_f3_runs_only_when_f1_and_f2_pass() -> None:
    scored = evaluate_f3(
        f1_status="PASS",
        f2_status="PASS",
        share_a=0.65,
        share_b=0.20,
        share_e=0.15,
        total_value_rub=1_000_000.0,
        entry_price_rub=500_000.0,
        largest_sku_share=0.20,
        match_confidence_weighted=0.9,
        core70_value_rub=800_000.0,
        share_used=0.1,
        used_signal_available=True,
        ticket_pct=0.5,
        core_sku_count=20,
        brand_count_core=2,
    )
    skipped = evaluate_f3(
        f1_status="STOP",
        f2_status="PASS",
        share_a=0.65,
        share_b=0.20,
        share_e=0.15,
        total_value_rub=1_000_000.0,
        entry_price_rub=500_000.0,
        largest_sku_share=0.20,
        match_confidence_weighted=0.9,
        core70_value_rub=800_000.0,
        share_used=0.1,
        used_signal_available=True,
        ticket_pct=0.5,
        core_sku_count=20,
        brand_count_core=2,
    )
    assert scored["score_100"] > 0
    assert skipped["score_100"] == 0


def test_json_artifact_is_deterministic_excluding_generated_at(tmp_path: Path) -> None:
    cfg = ProtocolConfig()
    payload = build_lot_artifact_payload(
        lot_id="1",
        statuses={"f0": "PASS", "f1": "PASS", "f2": "PASS", "f3": "SCORED"},
        metrics={"total_value_rub": 1_000_000.0},
        top3=[{"sku": "A", "raw_text": "A", "qty_total": 1.0, "sku_value_rub": 100_000.0, "category": "gas_safety", "brand": "honeywell"}],
        top20=[],
        top40=[],
        top80=[],
        review_lines=[],
        filter4_candidate=True,
        config=cfg,
    )
    p1 = write_lot_artifact(tmp_path, "1", payload)
    p2 = write_lot_artifact(tmp_path, "1_copy", payload)
    with p1.open("r", encoding="utf-8") as h1, p2.open("r", encoding="utf-8") as h2:
        left = json.load(h1)
        right = json.load(h2)
    left.pop("generated_at", None)
    right.pop("generated_at", None)
    assert left == right


def test_smoke_honeywell_optional(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[3]
    input_file = repo_root / "honeywell.xlsx"
    if not input_file.exists():
        pytest.skip("honeywell.xlsx not found")

    out_dir = tmp_path / "lots"
    result = run_filters(
        input_path=input_file,
        usd_rate=100.0,
        entry_price_rub=500_000.0,
        output_dir=out_dir,
        filter4_top_n=5,
    )
    assert Path(result["summary_path"]).exists()

    json_files = sorted(out_dir.glob("lot_*_artifact.json"))
    assert len(json_files) == 33
    assert (out_dir / "_metadata" / "engine_manifest.json").exists()
    assert (out_dir / "_metadata" / "merge_stats.json").exists()

    workbook = load_workbook(out_dir / "summary_filters_report.xlsx")
    ws = workbook.active
    header = [cell.value for cell in ws[1]]
    score_idx = header.index("score_100")
    f1_idx = header.index("f1_status")
    f2_idx = header.index("f2_status")
    for row in ws.iter_rows(min_row=2, values_only=True):
        score = row[score_idx] or 0
        f1_status = row[f1_idx]
        f2_status = row[f2_idx]
        if score > 0:
            assert f1_status == "PASS"
            assert f2_status == "PASS"
