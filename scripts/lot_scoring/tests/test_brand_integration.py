import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.lot_scoring.cdm import LotRecord
from scripts.lot_scoring.io.excel_report import write_investor_report
from scripts.lot_scoring.pipeline.brand_classifier import classify_lot_skus
from scripts.lot_scoring.pipeline.brand_metrics import compute_brand_flags, compute_brand_intelligence_summary, compute_cps
from scripts.lot_scoring.pipeline.explain_v4 import compute_v4_scalars
from scripts.lot_scoring.pipeline.intelligence_loader import load_all_brand_profiles
from scripts.lot_scoring.run_config import RunConfig
from scripts.lot_scoring.simulation.scoring_kernel import V4Config


def _profile_payload(brand: str) -> dict:
    return {
        "schema_version": "brand_intelligence_v1",
        "brand": brand,
        "metadata": {"author": "test"},
        "tuning_params": {
            "alpha_freshness_impact": 0.65,
            "cps_floor": 0.35,
            "reject_abs_usd_ceil": 50000.0,
            "legacy_combinator_mode": "max",
        },
        "verticals": {
            "hard_industrial": {
                "tier": "core",
                "purity_weight": 1.0,
                "obsolescence_score": 0.25,
                "freshness_score": 0.80,
                "reject": False,
            },
            "toxic_waste": {
                "tier": "reject",
                "purity_weight": 0.0,
                "obsolescence_score": 1.0,
                "freshness_score": 0.0,
                "reject": True,
            },
            "unknown_honeywell": {
                "tier": "unknown",
                "purity_weight": 0.2,
                "obsolescence_score": 0.50,
                "reject": False,
                "freshness_score_by_value": {
                    "high_threshold_usd": 500.0,
                    "high": 0.40,
                    "low": 0.20,
                },
            },
        },
        "category_mapping": {
            "keywords": {
                "relay": "hard_industrial",
                "obsolete": "toxic_waste",
            },
            "fallback_vertical": "unknown_honeywell",
        },
        "legacy_patterns": [
            {"pattern": "^ML1.*", "pattern_type": "regex", "penalty": 0.40, "reason": "legacy"},
        ],
        "classification_precedence": [
            "legacy_patterns",
            "category_mapping.keywords",
            "fallback_vertical",
        ],
    }


def _write_profile(tmp_path: Path, brand: str) -> None:
    target = tmp_path / f"{brand}.json"
    target.write_text(json.dumps(_profile_payload(brand), ensure_ascii=False), encoding="utf-8")


def test_excel_export_contains_brand_columns(tmp_path) -> None:
    rows = [
        {
            "rank": 1,
            "lot_id": "1",
            "Tier": "STRONG",
            "Zone_v4": "PASS",
            "FAS_v4": 70.0,
            "AQ_v4": 80.0,
            "score_10": 8.2,
            "abs_score_100": 72.0,
            "data_trust": "HIGH",
            "data_trust_pct": 0.8,
            "total_effective_usd": 1000.0,
            "hhi": 0.2,
            "unknown_exposure": 0.1,
            "confidence_score": 0.8,
            "lot_avg_pn_liquidity": 0.6,
            "needs_manual_review": False,
            "short_reason": "balanced_profile",
            "red_flags": "",
            "base_score": 0.7,
            "final_score": 0.68,
            "S1": 0.7,
            "S2": 0.6,
            "S3": 0.8,
            "S4": 0.5,
            "M_hustle": 0.9,
            "M_concentration": 0.95,
            "M_tail": 0.92,
            "M_Lifecycle": 0.9,
            "M_Unknown": 0.88,
            "KS_liquidity": 1.0,
            "KS_obsolescence": 1.0,
            "core_ratio": 0.75,
            "core_count": 2,
            "tail_count": 1,
            "unknown_exposure_usd": 100.0,
            "governance_flags": "",
            "flags": "",
            "Top_Loss_Reasons": "M_Unknown:30%",
            "Investor_Summary": "AQ=80.0, FAS=70.0",
            "brand_detected": "honeywell",
            "brand_cps": 0.42,
            "brand_flags": "LOW_CAPITAL_PURITY",
            "brand_verticals_json": "{\"honeywell\":{\"hard_industrial\":{\"usd\":1000.0,\"pct\":1.0}}}",
        }
    ]
    out = tmp_path / "report.xlsx"
    write_investor_report(rows, out)

    wb = load_workbook(out)
    ws = wb["RANKING"]
    headers = [str(cell.value) for cell in ws[1]]
    assert "brand_detected" in headers
    assert "brand_cps" in headers
    assert "brand_flags" in headers
    assert "brand_verticals_json" in headers


def test_multi_brand_classification_and_summary(tmp_path) -> None:
    _write_profile(tmp_path, "honeywell")
    _write_profile(tmp_path, "siemens")
    profiles = load_all_brand_profiles(base_dir=tmp_path)
    assert set(profiles.keys()) == {"honeywell", "siemens"}

    all_skus = [
        {
            "brand": "honeywell",
            "sku_code_normalized": "ML1-111",
            "description": "safety relay",
            "effective_line_usd": 1000.0,
            "q_has_qty": True,
        },
        {
            "brand": "siemens",
            "sku_code_normalized": "ML1-222",
            "description": "safety relay",
            "effective_line_usd": 1200.0,
            "q_has_qty": True,
        },
    ]
    classify_lot_skus(all_skus, profiles)

    honeywell_cps = compute_cps(all_skus, profiles["honeywell"])
    honeywell_flags = compute_brand_flags(honeywell_cps, profiles["honeywell"])
    honeywell_summary = compute_brand_intelligence_summary(all_skus, profiles["honeywell"], honeywell_cps, honeywell_flags)

    siemens_cps = compute_cps(all_skus, profiles["siemens"])
    siemens_flags = compute_brand_flags(siemens_cps, profiles["siemens"])
    siemens_summary = compute_brand_intelligence_summary(all_skus, profiles["siemens"], siemens_cps, siemens_flags)

    assert honeywell_summary["brand"] == "honeywell"
    assert siemens_summary["brand"] == "siemens"
    assert honeywell_summary["total_usd"] == 1000.0
    assert siemens_summary["total_usd"] == 1200.0


def test_low_cps_does_not_skip_fas_path(tmp_path) -> None:
    _write_profile(tmp_path, "honeywell")
    profiles = load_all_brand_profiles(base_dir=tmp_path)
    all_skus = [
        {
            "brand": "honeywell",
            "sku_code_normalized": "ML1-333",
            "description": "obsolete relay",
            "effective_line_usd": 100000.0,
            "q_has_qty": True,
            "category": "unknown",
            "is_in_core_slice": True,
        },
    ]
    classify_lot_skus(all_skus, profiles)
    cps_result = compute_cps(all_skus, profiles["honeywell"])
    flags = compute_brand_flags(cps_result, profiles["honeywell"])
    summary = compute_brand_intelligence_summary(all_skus, profiles["honeywell"], cps_result, flags)

    lot = LotRecord(lot_id="L-T")
    lot.s1 = 0.5
    lot.s2 = 0.6
    lot.s3 = 0.5
    lot.s4 = 0.5
    lot.hhi_index = 0.2
    lot.unknown_exposure = 0.1
    lot.penny_line_ratio = 0.1
    lot.core_distinct_sku_count = 1
    v4 = compute_v4_scalars(
        lot=lot,
        metadata={"core_count": 1, "tail_count": 0},
        core_skus=all_skus,
        category_obsolescence={"unknown": 0.4, "_default": 0.4},
        run_config=RunConfig(),
        v4_config=V4Config(),
    )

    assert "fas" in v4
    assert "aq" in v4
    assert summary["cps"] <= 1.0
