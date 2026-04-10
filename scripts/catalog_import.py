"""
R1.1 — Catalog Import Pipeline: Excel/CSV → stg_catalog_imports records.

Reads source Excel (honeywell_aggregated_named_v12.xlsx), validates input
contract, enriches from evidence bundles, applies confidence gate, and
routes each row to accepted / review_required / rejected.

Output: stg_catalog_imports.json + import_report.json (always, even --dry-run).

Usage:
    python scripts/catalog_import.py --dry-run
    python scripts/catalog_import.py --input path/to/file.xlsx
    python scripts/catalog_import.py --input path/to/file.csv
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
DOWNLOADS = ROOT / "downloads"
EVIDENCE_DIR = DOWNLOADS / "evidence"
OUTPUT_DIR = DOWNLOADS / "staging"

DEFAULT_INPUT = Path(ROOT).parent / "downloads" / "honeywell_aggregated_named_v12.xlsx"

# ── Constants (no magic strings) ────────────────────────────────────────────

# Confidence levels (from catalog_evidence_policy_v1)
CONFIDENCE_HIGH = "HIGH"
CONFIDENCE_MEDIUM = "MEDIUM"
CONFIDENCE_LOW = "LOW"

# Import statuses (aligned with migration 021 FSM: max 5 states)
STATUS_PENDING = "pending"
STATUS_ACCEPTED = "accepted"
STATUS_REVIEW_REQUIRED = "review_required"
STATUS_REJECTED = "rejected"

# Review reasons (from R1.7 spec)
REASON_VALIDATION_FAILED = "validation_failed"
REASON_NO_EVIDENCE = "no_evidence"
REASON_MISSING_PRICE = "missing_price"
REASON_NO_PHOTO = "no_photo"
REASON_TITLE_CONFIDENCE_LOW = "title_confidence_low"
REASON_DUPLICATE_PN = "duplicate_pn"

# Error classes (DNA §7)
ERROR_POLICY_VIOLATION = "POLICY_VIOLATION"
ERROR_PERMANENT = "PERMANENT"

# Excel column mapping
EXCEL_COLUMNS = {
    "part_number": "part_number",
    "brand": "brand",
    "quantity": "quantity",
    "unit_price": "unit_price",
    "title_ru": "title_ru",
    "source_name": "source_name",
    "condition_ru": "condition_ru",
}


# ── Parsing ─────────────────────────────────────────────────────────────────

def parse_excel(path: Path) -> list[dict[str, Any]]:
    """Parse Excel (.xlsx) file into list of row dicts. Normalizes keys."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    # Use first sheet by default
    ws = wb[wb.sheetnames[0]]

    headers = []
    rows: list[dict[str, Any]] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip().lower() if h else f"col_{j}" for j, h in enumerate(row)]
            continue
        if all(c is None for c in row):
            continue
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)

    wb.close()
    return rows


def parse_csv(path: Path) -> list[dict[str, Any]]:
    """Parse CSV file into list of row dicts."""
    rows: list[dict[str, Any]] = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k.strip().lower(): v for k, v in row.items()})
    return rows


def parse_input(path: Path) -> list[dict[str, Any]]:
    """Parse input file (xlsx or csv) into list of row dicts."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return parse_excel(path)
    if suffix == ".csv":
        return parse_csv(path)
    raise ValueError(f"Unsupported file format: {suffix} (expected .xlsx or .csv)")


# ── Normalization ───────────────────────────────────────────────────────────

def normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Normalize a raw parsed row: strip, uppercase PN/brand, coerce types."""
    pn = str(raw.get("part_number") or "").strip().upper()
    brand = str(raw.get("brand") or "").strip().upper()

    # Quantity: coerce to int
    qty_raw = raw.get("quantity")
    try:
        qty = int(float(qty_raw)) if qty_raw is not None else 0
    except (ValueError, TypeError):
        qty = 0

    # Price: coerce to float
    price_raw = raw.get("unit_price")
    try:
        price = float(price_raw) if price_raw is not None else None
    except (ValueError, TypeError):
        price = None

    return {
        "part_number": pn,
        "brand": brand,
        "quantity": qty,
        "unit_price": price,
        "title_ru": str(raw.get("title_ru") or "").strip(),
        "source_name": str(raw.get("source_name") or "").strip(),
        "condition_ru": str(raw.get("condition_ru") or "").strip(),
    }


# ── Validation ──────────────────────────────────────────────────────────────

def validate_row(row: dict[str, Any]) -> tuple[bool, str | None]:
    """Validate input contract. Returns (valid, reject_reason).

    Required: part_number (non-empty), brand (non-empty), quantity (>0).
    unit_price is optional.
    """
    if not row.get("part_number"):
        return False, "missing_part_number"
    if not row.get("brand"):
        return False, "missing_brand"
    qty = row.get("quantity", 0)
    if not isinstance(qty, int) or qty <= 0:
        return False, "missing_or_zero_quantity"
    return True, None


# ── Evidence enrichment ─────────────────────────────────────────────────────

def _load_evidence(pn: str, evidence_dir: Path) -> dict | None:
    """Load evidence bundle for a PN. Returns None if not found."""
    evidence_path = evidence_dir / f"evidence_{pn}.json"
    if not evidence_path.exists():
        return None
    try:
        return json.loads(evidence_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def enrich_from_evidence(
    row: dict[str, Any], evidence_dir: Path,
) -> dict[str, Any]:
    """Enrich a normalized row with data from its evidence bundle.

    Adds: confidence, photo_url, review_reasons, evidence_found.
    """
    pn = row["part_number"]
    evidence = _load_evidence(pn, evidence_dir)

    if evidence is None:
        row["confidence"] = CONFIDENCE_LOW
        row["photo_url"] = None
        row["review_reasons"] = [REASON_NO_EVIDENCE]
        row["evidence_found"] = False
        return row

    row["evidence_found"] = True

    # Confidence from evidence bundle
    conf = evidence.get("confidence", {})
    overall_label = conf.get("overall_label", "").upper()
    if overall_label in (CONFIDENCE_HIGH, CONFIDENCE_MEDIUM, CONFIDENCE_LOW):
        row["confidence"] = overall_label
    else:
        row["confidence"] = CONFIDENCE_LOW

    # Photo
    photo = evidence.get("photo", {})
    photo_verdict = photo.get("verdict", "").upper()
    row["photo_url"] = photo.get("photo_url") if photo_verdict == "ACCEPT" else None

    # Collect review reasons from evidence
    review_reasons: list[str] = []

    # Price check
    price_data = evidence.get("price", {})
    price_status = price_data.get("price_status", "")
    if not row.get("unit_price") and price_status != "public_price":
        review_reasons.append(REASON_MISSING_PRICE)

    # Photo check
    if photo_verdict != "ACCEPT":
        review_reasons.append(REASON_NO_PHOTO)

    # Identity check
    identity_level = evidence.get("identity_level", "")
    if identity_level in ("weak", "unknown"):
        review_reasons.append(REASON_TITLE_CONFIDENCE_LOW)

    row["review_reasons"] = review_reasons
    return row


# ── Classification ──────────────────────────────────────────────────────────

def classify_row(row: dict[str, Any]) -> dict[str, Any]:
    """Assign import status based on confidence + review reasons.

    Routing:
        HIGH confidence + no blocking reasons → accepted
        MEDIUM confidence → review_required
        LOW confidence → review_required (with reasons)
        Validation failures → rejected
    """
    confidence = row.get("confidence", CONFIDENCE_LOW)
    review_reasons = row.get("review_reasons", [])

    if confidence == CONFIDENCE_HIGH and not review_reasons:
        row["status"] = STATUS_ACCEPTED
        row["review_reason"] = None
    elif confidence == CONFIDENCE_HIGH and review_reasons:
        # HIGH but has issues → review
        row["status"] = STATUS_REVIEW_REQUIRED
        row["review_reason"] = review_reasons[0]
    elif confidence == CONFIDENCE_MEDIUM:
        row["status"] = STATUS_REVIEW_REQUIRED
        row["review_reason"] = review_reasons[0] if review_reasons else "medium_confidence"
    else:
        # LOW
        row["status"] = STATUS_REVIEW_REQUIRED
        row["review_reason"] = review_reasons[0] if review_reasons else REASON_NO_EVIDENCE

    return row


# ── Import record builder ──────────────────────────────────────────────────

def build_import_record(
    row: dict[str, Any],
    job_id: str,
    trace_id: str,
) -> dict[str, Any]:
    """Build a stg_catalog_imports-compatible record (migration 021 schema)."""
    pn = row["part_number"]
    brand = row["brand"]

    # Idempotency key: after normalization (owner requirement #5)
    idem_raw = f"{brand}:{pn}"
    idempotency_key = hashlib.sha256(idem_raw.encode("utf-8")).hexdigest()[:16]

    return {
        "id": str(uuid.uuid4()),
        "job_id": job_id,
        "trace_id": trace_id,
        "idempotency_key": idempotency_key,
        "brand": brand,
        "part_number": pn,
        "name": row.get("title_ru", ""),
        "qty": row.get("quantity", 0),
        "approx_price": row.get("unit_price"),
        "confidence": row.get("confidence", CONFIDENCE_LOW),
        "review_reason": row.get("review_reason"),
        "photo_url": row.get("photo_url"),
        "status": row.get("status", STATUS_PENDING),
        "error_class": row.get("error_class"),
        "error": row.get("error"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }


# ── Main pipeline ───────────────────────────────────────────────────────────

def run_import(
    input_path: Path,
    evidence_dir: Path = EVIDENCE_DIR,
    output_dir: Path = OUTPUT_DIR,
    *,
    trace_id: str = "",
    dry_run: bool = False,
) -> dict[str, Any]:
    """Run the full R1.1 import pipeline.

    Returns import_report dict. Always writes import_report.json (even dry-run).
    """
    if not trace_id:
        trace_id = f"r1-import-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}"

    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    # Parse input
    raw_rows = parse_input(input_path)
    logger.info("Parsed %d rows from %s", len(raw_rows), input_path.name)

    # Normalize all rows
    normalized = [normalize_row(r) for r in raw_rows]

    # Validate + dedup (after normalization — owner requirement #5)
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    seen_keys: set[str] = set()

    for row in normalized:
        # Validation
        valid, reject_reason = validate_row(row)
        if not valid:
            row["status"] = STATUS_REJECTED
            row["error_class"] = ERROR_POLICY_VIOLATION
            row["error"] = reject_reason
            row["review_reason"] = REASON_VALIDATION_FAILED
            rejected.append(build_import_record(row, job_id, trace_id))
            continue

        # Dedup by brand:pn
        dedup_key = f"{row['brand']}:{row['part_number']}"
        if dedup_key in seen_keys:
            row["status"] = STATUS_REJECTED
            row["error_class"] = ERROR_PERMANENT
            row["error"] = "duplicate_in_batch"
            row["review_reason"] = REASON_DUPLICATE_PN
            rejected.append(build_import_record(row, job_id, trace_id))
            continue
        seen_keys.add(dedup_key)

        # Enrich from evidence
        row = enrich_from_evidence(row, evidence_dir)

        # Classify
        row = classify_row(row)

        records.append(build_import_record(row, job_id, trace_id))

    # Split by status
    accepted = [r for r in records if r["status"] == STATUS_ACCEPTED]
    review = [r for r in records if r["status"] == STATUS_REVIEW_REQUIRED]
    all_records = records + rejected

    # Build report
    reason_counts: dict[str, int] = {}
    for r in all_records:
        reason = r.get("review_reason") or "none"
        reason_counts[reason] = reason_counts.get(reason, 0) + 1

    brand_counts: dict[str, dict[str, int]] = {}
    for r in all_records:
        b = r["brand"]
        s = r["status"]
        if b not in brand_counts:
            brand_counts[b] = {}
        brand_counts[b][s] = brand_counts[b].get(s, 0) + 1

    report = {
        "trace_id": trace_id,
        "job_id": job_id,
        "input_file": str(input_path),
        "dry_run": dry_run,
        "timestamp": now,
        "total_parsed": len(raw_rows),
        "total_after_dedup": len(records),
        "accepted": len(accepted),
        "review_required": len(review),
        "rejected": len(rejected),
        "by_reason": dict(sorted(reason_counts.items())),
        "by_brand": dict(sorted(brand_counts.items())),
        "confidence_distribution": {
            CONFIDENCE_HIGH: sum(1 for r in records if r["confidence"] == CONFIDENCE_HIGH),
            CONFIDENCE_MEDIUM: sum(1 for r in records if r["confidence"] == CONFIDENCE_MEDIUM),
            CONFIDENCE_LOW: sum(1 for r in records if r["confidence"] == CONFIDENCE_LOW),
        },
    }

    # Write outputs (always — even dry-run, per owner requirement #2)
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "import_report.json"
    report_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8",
    )
    logger.info("Report written: %s", report_path)

    if not dry_run:
        records_path = output_dir / "stg_catalog_imports.json"
        records_path.write_text(
            json.dumps(all_records, indent=2, ensure_ascii=False), encoding="utf-8",
        )
        logger.info("Records written: %s (%d records)", records_path, len(all_records))

    # Console summary
    print(f"\n{'='*60}")
    print(f"R1.1 Import Pipeline {'(DRY RUN)' if dry_run else ''}")
    print(f"{'='*60}")
    print(f"Input:            {input_path.name}")
    print(f"Total parsed:     {report['total_parsed']}")
    print(f"After dedup:      {report['total_after_dedup']}")
    print(f"  Accepted:       {report['accepted']}")
    print(f"  Review required:{report['review_required']}")
    print(f"  Rejected:       {report['rejected']}")
    print(f"\nConfidence: HIGH={report['confidence_distribution']['HIGH']} "
          f"MEDIUM={report['confidence_distribution']['MEDIUM']} "
          f"LOW={report['confidence_distribution']['LOW']}")
    print("\nTop reasons:")
    for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1])[:7]:
        print(f"  {reason}: {count}")
    print(f"\nReport: {report_path}")
    if not dry_run:
        print(f"Records: {output_dir / 'stg_catalog_imports.json'}")
    print(f"{'='*60}\n")

    return report


# ── CLI ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="R1.1 Catalog Import Pipeline — Excel/CSV → stg_catalog_imports",
    )
    parser.add_argument(
        "--input", type=Path, default=DEFAULT_INPUT,
        help="Input Excel (.xlsx) or CSV file",
    )
    parser.add_argument(
        "--evidence-dir", type=Path, default=EVIDENCE_DIR,
        help="Evidence bundles directory",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=OUTPUT_DIR,
        help="Output directory for staging records",
    )
    parser.add_argument(
        "--trace-id", default="",
        help="Trace ID for this import run",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse and validate only, write report but not staging records",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if not args.input.exists():
        print(f"ERROR: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    run_import(
        input_path=args.input,
        evidence_dir=args.evidence_dir,
        output_dir=args.output_dir,
        trace_id=args.trace_id,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
