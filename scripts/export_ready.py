#!/usr/bin/env python3
"""
export_ready.py — EXPORT-READY CONTROL LAYER v1

Reads all evidence files, computes a FRESH readiness view based on
current DR-enriched data (NOT stale card_status from photo_pipeline runs).

Key insight: card_status in evidence is stale (computed before DR enrichment).
This script evaluates readiness from normalized{} block (evidence_normalize.py),
which unifies price/description/photo from all 3 pipelines into canonical fields.

Outputs:
  1. downloads/export/export_ready_view.json   — per-SKU readiness record
  2. downloads/export/draft_insales_export.xlsx — multi-sheet draft for InSales
  3. downloads/export/photo_manifest.csv        — photo source + cloud mapping
  4. downloads/export/missing_data_queue.csv    — gaps queue for next batch

Export-readiness status values:
  EXPORT_READY   — dr_price validated + no CRITICAL_MISMATCH → ready for InSales
  DRAFT_EXPORT   — no dr_price but our_price_raw exists + no CRITICAL_MISMATCH
  REVIEW_BLOCKED — CRITICAL_MISMATCH present (identity uncertain)
  BLOCKED_NO_PRICE — CRITICAL_MISMATCH + no price at all

Usage:
    python scripts/export_ready.py [--dry-run] [--pn PN]
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
EVIDENCE_DIR = ROOT / "downloads" / "evidence"
PHOTOS_DIR = ROOT / "downloads" / "photos"
EXPORT_DIR = ROOT / "downloads" / "export"

# ── FX approximate rates (stub — live FX not yet implemented, see fx.py P0-1) ──
_TO_RUB_APPROX: dict[str, float] = {
    "RUB": 1.0,
    "USD": 90.0,
    "EUR": 97.0,
    "GBP": 114.0,
    "CHF": 101.0,
    "CZK": 3.8,
    "PLN": 22.5,
    "DKK": 13.0,
    "SEK": 8.4,
    "NOK": 8.2,
    "HUF": 0.24,
    "ARS": 0.10,  # volatile, flag separately
    "KZT": 0.19,
    "JPY": 0.60,
    "TWD": 2.8,
}

_VOLATILE_CURRENCIES = {"ARS", "KZT", "JPY", "TWD"}

# ── Blocking identity reason codes ──────────────────────────────────────────────
_IDENTITY_BLOCKERS = {"CRITICAL_MISMATCH", "mismatch_blocks_publish"}


def _is_empty(val) -> bool:
    if val is None:
        return True
    if isinstance(val, str):
        return val.strip() in ("", "not_found", "not found", "none", "null", "-----------")
    if isinstance(val, (dict, list)):
        return len(val) == 0
    return False


def _to_rub(amount: float, currency: str) -> tuple[float | None, bool]:
    """Convert to RUB using stub rates. Returns (rub_amount, is_volatile)."""
    cur = (currency or "USD").upper()
    rate = _TO_RUB_APPROX.get(cur)
    if rate is None:
        return None, False
    return round(amount * rate, 2), cur in _VOLATILE_CURRENCIES


def _has_identity_blocker(review_reasons: list) -> bool:
    for r in review_reasons:
        code = r.get("code", r) if isinstance(r, dict) else str(r)
        if code in _IDENTITY_BLOCKERS:
            return True
    return False


def _local_photo_path(pn: str) -> str | None:
    """Return path to local photo file if it exists."""
    for ext in ("jpg", "jpeg", "png", "webp"):
        p = PHOTOS_DIR / f"{pn}.{ext}"
        if p.exists():
            return str(p)
    # Fuzzy: sanitized PN
    safe_pn = pn.replace("/", "_").replace("\\", "_")
    for ext in ("jpg", "jpeg", "png", "webp"):
        p = PHOTOS_DIR / f"{safe_pn}.{ext}"
        if p.exists():
            return str(p)
    return None


def _suggested_cloud_key(brand: str, pn: str) -> str:
    """Suggest cloud storage object key for photo upload."""
    safe_brand = (brand or "unknown").upper().replace(" ", "_")[:20]
    safe_pn = pn.replace("/", "_").replace(".", "_").replace("\\", "_")
    return f"catalog/{safe_brand}/{safe_pn}.jpg"


def compute_readiness(evidence: dict) -> dict:
    """
    Compute fresh export-readiness record from evidence fields.

    Does NOT use stale card_status from photo_pipeline runs.
    Reads price/description/photo from normalized{} block (evidence_normalize.py).
    Falls back to raw evidence fields when normalized{} is absent.
    """
    pn = evidence.get("pn", "")
    brand = evidence.get("brand", "Honeywell") or "Honeywell"
    subbrand = evidence.get("subbrand", "") or ""
    assembled_title = evidence.get("assembled_title", "") or ""
    product_category = evidence.get("product_category", "") or ""
    review_reasons_raw = evidence.get("review_reasons", []) or []

    dr = evidence.get("deep_research", {}) or {}
    title_ru = dr.get("title_ru", "") or ""
    specs = dr.get("specs", {}) or {}

    # dr_price_blocked still read directly — it's a quality gate, not a data field
    dr_price_blocked = evidence.get("dr_price_blocked")
    dr_price_source = evidence.get("dr_price_source", "") or ""

    # ── Normalized block (single source of truth for price/desc/photo) ───────────
    norm = evidence.get("normalized") or {}
    best_price: float | None = norm.get("best_price")
    best_price_currency: str = norm.get("best_price_currency") or ""
    best_price_source: str | None = norm.get("best_price_source")  # price_contract|pipeline1|our_estimate
    best_description: str = (norm.get("best_description") or "").strip()
    best_photo_url: str = (norm.get("best_photo_url") or "").strip()

    # ── Title ────────────────────────────────────────────────────────────────────
    display_title = title_ru.strip() if not _is_empty(title_ru) else assembled_title.strip()
    title_ok = not _is_empty(display_title)
    title_status = "ok" if title_ok else "missing"

    # ── Price (from normalized, quality-gated by dr_price_blocked) ───────────────
    # Market sources: price_contract (DR pipeline) + pipeline1 (Phase A SerpAPI)
    # Estimate source: our_estimate (RUB market estimate from internal Excel)
    price_market = (
        best_price is not None
        and best_price_source in ("price_contract", "pipeline1")
        and not dr_price_blocked
    )
    price_estimate = (
        best_price is not None
        and best_price_source == "our_estimate"
        and not price_market
    )

    export_price_rub: float | None = None
    price_source_label = "none"
    price_currency_original = ""
    price_volatile = False

    if price_market:
        rub, is_vol = _to_rub(float(best_price), best_price_currency)
        export_price_rub = rub
        price_source_label = best_price_source  # "price_contract" or "pipeline1"
        price_currency_original = best_price_currency
        price_volatile = is_vol
    elif price_estimate:
        export_price_rub = float(best_price)  # already in RUB
        price_source_label = "our_estimate"
        price_currency_original = "RUB"

    if price_market:
        price_status = "validated_market"
    elif price_estimate:
        price_status = "ref_price_only"
    elif dr_price_blocked:
        price_status = "blocked"
        flags = dr_price_blocked.get("flags", []) if isinstance(dr_price_blocked, dict) else []
        if "DR_PACK_SIGNAL_DETECTED" in flags:
            price_status = "blocked_pack_signal"
    else:
        price_status = "missing"

    # ── Photo (from normalized — respects photo.verdict KEEP/ACCEPT) ─────────────
    photo_url = best_photo_url
    local_path = _local_photo_path(pn) if not photo_url else None

    if photo_url:
        photo_status = "url_available"
        photo_asset = photo_url
        photo_ready_for_cloud = True
    elif local_path:
        photo_status = "local_file"
        photo_asset = local_path
        photo_ready_for_cloud = True
    else:
        photo_status = "missing"
        photo_asset = ""
        photo_ready_for_cloud = False

    # ── Identity ─────────────────────────────────────────────────────────────────
    identity_blocked = _has_identity_blocker(review_reasons_raw)
    identity_status = "blocked" if identity_blocked else "ok"

    # ── Specs ────────────────────────────────────────────────────────────────────
    specs_status = "ok" if not _is_empty(specs) else "missing"

    # ── Description (from normalized — best of dr.description_ru + content.description) ──
    desc_status = "ok" if len(best_description) >= 10 else "missing"

    # ── Category ─────────────────────────────────────────────────────────────────
    category_status = "ok" if not _is_empty(product_category) else "missing"

    # ── Export readiness ─────────────────────────────────────────────────────────
    # Blocking rules:
    # 1. identity_blocked → always blocks
    # 2. price_ok_dr required for EXPORT_READY (not just ref price)
    # 3. title always available (100%), so never a blocker in practice
    missing_fields = []
    if not title_ok:
        missing_fields.append("title")
    if price_status == "missing":
        missing_fields.append("price")
    if price_status in ("blocked", "blocked_pack_signal"):
        missing_fields.append("price_blocked")
    if photo_status == "missing":
        missing_fields.append("photo")
    if desc_status == "missing":
        missing_fields.append("description")
    if specs_status == "missing":
        missing_fields.append("specs")
    if identity_blocked:
        missing_fields.append("identity_conflict")

    review_needed = identity_blocked or price_volatile or price_source_label == "our_estimate"

    if identity_blocked:
        if price_market or price_estimate:
            export_status = "REVIEW_BLOCKED"
        else:
            export_status = "BLOCKED_NO_PRICE"
    elif price_market:
        export_status = "EXPORT_READY"
    elif price_estimate:
        export_status = "DRAFT_EXPORT"
    else:
        export_status = "BLOCKED_NO_PRICE"

    # ── Missing-data severity ────────────────────────────────────────────────────
    best_next_action = ""
    if export_status == "REVIEW_BLOCKED":
        best_next_action = f"Resolve identity conflict: re-run DR with subbrand hint for {pn}"
    elif export_status == "BLOCKED_NO_PRICE":
        best_next_action = f"Find price: DR or Gemini fast for {pn} ({product_category or 'unknown category'})"
    elif export_status == "DRAFT_EXPORT":
        best_next_action = f"Get market price: DR/Gemini fast for {pn} to replace ref price"
    elif price_volatile:
        best_next_action = f"Verify price: {best_price_currency} is volatile currency, check manually"

    return {
        "pn": pn,
        "brand": brand,
        "subbrand": subbrand,
        "display_title": display_title,
        "product_category": product_category,
        "export_status": export_status,
        "export_price_rub": export_price_rub,
        "price_currency_original": price_currency_original,
        "price_source": price_source_label,
        "price_status": price_status,
        "price_volatile": price_volatile,
        "dr_price_source_url": dr_price_source,
        "photo_status": photo_status,
        "photo_asset": photo_asset,
        "photo_ready_for_cloud": photo_ready_for_cloud,
        "cloud_key": _suggested_cloud_key(brand, pn),
        "title_status": title_status,
        "desc_status": desc_status,
        "specs_status": specs_status,
        "category_status": category_status,
        "identity_status": identity_status,
        "review_needed": review_needed,
        "missing_fields": missing_fields,
        "best_next_action": best_next_action,
        "description_ru": best_description[:300] if best_description else "",
    }


def build_readiness_view(pn_filter: str | None = None) -> list[dict]:
    """Load all evidence files and compute readiness for each SKU."""
    evidence_files = sorted(EVIDENCE_DIR.glob("evidence_*.json"))
    records = []

    for ev_path in evidence_files:
        try:
            evidence = json.loads(ev_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        pn = evidence.get("pn", "")
        if not pn:
            continue
        if pn_filter and pn != pn_filter:
            continue

        rec = compute_readiness(evidence)
        records.append(rec)

    return records


def write_export_view_json(records: list[dict], dry_run: bool = False) -> Path:
    """Write per-SKU readiness view to JSON."""
    out_path = EXPORT_DIR / "export_ready_view.json"
    payload = {
        "generated_ts": datetime.now(timezone.utc).isoformat(),
        "total": len(records),
        "summary": {
            s: sum(1 for r in records if r["export_status"] == s)
            for s in ("EXPORT_READY", "DRAFT_EXPORT", "REVIEW_BLOCKED", "BLOCKED_NO_PRICE")
        },
        "skus": records,
    }
    if not dry_run:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return out_path


def write_insales_excel(records: list[dict], dry_run: bool = False) -> Path:
    """Write multi-sheet draft InSales Excel export."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    out_path = EXPORT_DIR / "draft_insales_export.xlsx"
    wb = openpyxl.Workbook()

    # ── Color scheme ─────────────────────────────────────────────────────────────
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    ORANGE = PatternFill("solid", fgColor="FFCC99")
    RED = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F497D")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    STATUS_COLOR = {
        "EXPORT_READY": GREEN,
        "DRAFT_EXPORT": YELLOW,
        "REVIEW_BLOCKED": ORANGE,
        "BLOCKED_NO_PRICE": RED,
    }

    # ── Column definitions ────────────────────────────────────────────────────────
    COLS_MAIN = [
        ("Артикул", "pn"),
        ("Название", "display_title"),
        ("Бренд", "brand"),
        ("Категория", "product_category"),
        ("Цена (RUB)", "export_price_rub"),
        ("Цена оригинал", "price_currency_original"),
        ("Источник цены", "price_source"),
        ("Статус цены", "price_status"),
        ("Фото", "photo_asset"),
        ("Статус фото", "photo_status"),
        ("Описание", "description_ru"),
        ("Статус описания", "desc_status"),
        ("Статус экспорта", "export_status"),
        ("Требует проверки", "review_needed"),
        ("Отсутствующие поля", "missing_fields"),
        ("Рекомендация", "best_next_action"),
    ]

    def add_sheet(wb, name: str, rows: list[dict], cols: list, status_col_idx: int = 12):
        ws = wb.create_sheet(title=name)
        # Header
        for ci, (label, _) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        # Data
        for ri, rec in enumerate(rows, 2):
            status = rec.get("export_status", "")
            fill = STATUS_COLOR.get(status)
            for ci, (_, key) in enumerate(cols, 1):
                val = rec.get(key, "")
                if isinstance(val, list):
                    val = "; ".join(val)
                elif isinstance(val, bool):
                    val = "Да" if val else "Нет"
                elif val is None:
                    val = ""
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
        # Column widths
        col_widths = [15, 50, 15, 30, 14, 12, 18, 20, 60, 14, 80, 16, 18, 14, 40, 60]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
        ws.freeze_panes = "A2"
        return ws

    # ── Sheet 0: Summary ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    summary_rows = [
        ["EXPORT-READY CONTROL LAYER v1", "", ""],
        ["Дата генерации", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"), ""],
        ["", "", ""],
        ["Статус", "Кол-во SKU", "Описание"],
        ["EXPORT_READY", sum(1 for r in records if r["export_status"] == "EXPORT_READY"),
         "Рыночная цена + без конфликта идентичности → готов к загрузке в InSales"],
        ["DRAFT_EXPORT", sum(1 for r in records if r["export_status"] == "DRAFT_EXPORT"),
         "Оценочная цена (our_estimate, RUB ~2022) → нужна рыночная цена"],
        ["REVIEW_BLOCKED", sum(1 for r in records if r["export_status"] == "REVIEW_BLOCKED"),
         "CRITICAL_MISMATCH: конфликт идентичности → нельзя публиковать"],
        ["BLOCKED_NO_PRICE", sum(1 for r in records if r["export_status"] == "BLOCKED_NO_PRICE"),
         "Нет цены + проблема идентичности"],
        ["", "", ""],
        ["ИТОГО SKU", len(records), ""],
        ["", "", ""],
        ["Покрытие полей", "Кол-во", "%"],
        ["Цена (рыночная)", sum(1 for r in records if r["price_source"] in ("price_contract", "pipeline1")),
         f"{int(100*sum(1 for r in records if r['price_source'] in ('price_contract','pipeline1'))/len(records))}%"],
        ["Цена (оценочная)", sum(1 for r in records if r["price_source"] == "our_estimate"),
         f"{int(100*sum(1 for r in records if r['price_source']=='our_estimate')/len(records))}%"],
        ["Фото (url/local)", sum(1 for r in records if r["photo_status"] != "missing"),
         f"{int(100*sum(1 for r in records if r['photo_status']!='missing')/len(records))}%"],
        ["Описание", sum(1 for r in records if r["desc_status"] == "ok"),
         f"{int(100*sum(1 for r in records if r['desc_status']=='ok')/len(records))}%"],
        ["Категория", sum(1 for r in records if r["category_status"] == "ok"),
         f"{int(100*sum(1 for r in records if r['category_status']=='ok')/len(records))}%"],
        ["Спецификации", sum(1 for r in records if r["specs_status"] == "ok"),
         f"{int(100*sum(1 for r in records if r['specs_status']=='ok')/len(records))}%"],
    ]
    for ri, row in enumerate(summary_rows, 1):
        for ci, val in enumerate(row, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            if ri == 4 or ri == 12:
                cell.font = Font(bold=True)
            if ri in (5, 6, 7, 8):
                status = row[0]
                fill = STATUS_COLOR.get(status)
                if fill:
                    cell.fill = fill
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 70

    # ── Sheet 1: EXPORT_READY ─────────────────────────────────────────────────────
    export_ready = [r for r in records if r["export_status"] == "EXPORT_READY"]
    add_sheet(wb, f"EXPORT_READY ({len(export_ready)})", export_ready, COLS_MAIN)

    # ── Sheet 2: DRAFT_EXPORT ─────────────────────────────────────────────────────
    draft = [r for r in records if r["export_status"] == "DRAFT_EXPORT"]
    add_sheet(wb, f"DRAFT_EXPORT ({len(draft)})", draft, COLS_MAIN)

    # ── Sheet 3: REVIEW_BLOCKED ───────────────────────────────────────────────────
    blocked = [r for r in records if r["export_status"] in ("REVIEW_BLOCKED", "BLOCKED_NO_PRICE")]
    add_sheet(wb, f"BLOCKED ({len(blocked)})", blocked, COLS_MAIN)

    if not dry_run:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
    return out_path


def write_photo_manifest(records: list[dict], dry_run: bool = False) -> Path:
    """Write photo manifest CSV for cloud upload planning."""
    out_path = EXPORT_DIR / "photo_manifest.csv"
    fieldnames = [
        "pn", "brand", "photo_status", "photo_asset",
        "ready_for_cloud", "suggested_cloud_key",
        "public_url_after_upload", "export_photo_field",
    ]
    rows = []
    for r in records:
        rows.append({
            "pn": r["pn"],
            "brand": r["brand"],
            "photo_status": r["photo_status"],
            "photo_asset": r["photo_asset"],
            "ready_for_cloud": "yes" if r["photo_ready_for_cloud"] else "no",
            "suggested_cloud_key": r["cloud_key"],
            "public_url_after_upload": "",
            "export_photo_field": r["photo_asset"],
        })

    if not dry_run:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    return out_path


def write_missing_data_queue(records: list[dict], dry_run: bool = False) -> Path:
    """Write missing-data queue CSV for next batch planning."""
    out_path = EXPORT_DIR / "missing_data_queue.csv"
    fieldnames = [
        "pn", "brand", "product_category", "export_status",
        "gap_type", "severity", "current_value", "best_next_action",
    ]

    SEVERITY = {
        "price": "P0",       # blocks export
        "price_blocked": "P0",
        "identity_conflict": "P0",  # blocks export
        "photo": "P1",       # doesn't block but needed
        "title": "P0",
        "description": "P2",
        "specs": "P2",
        "category": "P2",
    }

    rows = []
    for r in records:
        pn = r["pn"]
        brand = r["brand"]
        cat = r["product_category"]
        status = r["export_status"]
        action = r["best_next_action"]

        for gap in r["missing_fields"]:
            sev = SEVERITY.get(gap, "P2")
            current = ""
            if gap == "price":
                current = "no normalized.best_price"
            elif gap == "price_blocked":
                current = f"blocked: {r.get('price_status','')}"
            elif gap == "identity_conflict":
                current = "CRITICAL_MISMATCH in review_reasons"
            elif gap == "photo":
                current = "no dr_image_url, no local file"
            elif gap == "description":
                current = "deep_research.description_ru empty"
            elif gap == "specs":
                current = "deep_research.specs empty"
            elif gap == "category":
                current = "product_category empty"

            rows.append({
                "pn": pn,
                "brand": brand,
                "product_category": cat,
                "export_status": status,
                "gap_type": gap,
                "severity": sev,
                "current_value": current,
                "best_next_action": action,
            })

    # Sort by severity then pn
    sev_order = {"P0": 0, "P1": 1, "P2": 2}
    rows.sort(key=lambda x: (sev_order.get(x["severity"], 3), x["pn"]))

    if not dry_run:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    return out_path


def run(dry_run: bool = False, pn_filter: str | None = None) -> dict:
    """Main entry point. Returns summary dict."""
    print(f"Loading evidence files from {EVIDENCE_DIR}...")
    records = build_readiness_view(pn_filter=pn_filter)
    print(f"Loaded {len(records)} SKUs")

    summary = {
        s: sum(1 for r in records if r["export_status"] == s)
        for s in ("EXPORT_READY", "DRAFT_EXPORT", "REVIEW_BLOCKED", "BLOCKED_NO_PRICE")
    }
    summary["total"] = len(records)

    # Field coverage
    summary["coverage"] = {
        "price_contract": sum(1 for r in records if r["price_source"] == "price_contract"),
        "price_pipeline1": sum(1 for r in records if r["price_source"] == "pipeline1"),
        "price_estimate": sum(1 for r in records if r["price_source"] == "our_estimate"),
        "photo_url": sum(1 for r in records if r["photo_status"] == "url_available"),
        "photo_local": sum(1 for r in records if r["photo_status"] == "local_file"),
        "photo_missing": sum(1 for r in records if r["photo_status"] == "missing"),
        "desc_ok": sum(1 for r in records if r["desc_status"] == "ok"),
        "specs_ok": sum(1 for r in records if r["specs_status"] == "ok"),
        "category_ok": sum(1 for r in records if r["category_status"] == "ok"),
    }

    print("\nExport-readiness breakdown:")
    print(f"  EXPORT_READY:     {summary['EXPORT_READY']:3d}  (market price + clean identity)")
    print(f"  DRAFT_EXPORT:     {summary['DRAFT_EXPORT']:3d}  (ref price only, needs market price)")
    print(f"  REVIEW_BLOCKED:   {summary['REVIEW_BLOCKED']:3d}  (CRITICAL_MISMATCH)")
    print(f"  BLOCKED_NO_PRICE: {summary['BLOCKED_NO_PRICE']:3d}  (no price + identity issues)")
    print(f"  TOTAL:            {summary['total']:3d}")
    print("\nField coverage:")
    cov = summary["coverage"]
    n = summary["total"]
    price_total = cov['price_contract'] + cov['price_pipeline1'] + cov['price_estimate']
    print(f"  Price (total):    {price_total:3d} / {n} ({int(100*price_total/n)}%)")
    print(f"    price_contract: {cov['price_contract']:3d}  (DR pipeline)")
    print(f"    pipeline1:      {cov['price_pipeline1']:3d}  (SerpAPI Phase A)")
    print(f"    our_estimate:   {cov['price_estimate']:3d}  (RUB estimate, ~2022)")
    print(f"  Photo (url):      {cov['photo_url']:3d} / {n}")
    print(f"  Photo (local):    {cov['photo_local']:3d} / {n}")
    print(f"  Photo (missing):  {cov['photo_missing']:3d} / {n}")
    print(f"  Description:      {cov['desc_ok']:3d} / {n} ({int(100*cov['desc_ok']/n)}%)")
    print(f"  Specs:            {cov['specs_ok']:3d} / {n} ({int(100*cov['specs_ok']/n)}%)")
    print(f"  Category:         {cov['category_ok']:3d} / {n} ({int(100*cov['category_ok']/n)}%)")

    # Write artifacts
    paths = {}

    print(f"\nWriting artifacts (dry_run={dry_run})...")

    p = write_export_view_json(records, dry_run=dry_run)
    paths["export_view_json"] = str(p)
    print(f"  1. {p}")

    p = write_insales_excel(records, dry_run=dry_run)
    paths["insales_excel"] = str(p)
    print(f"  2. {p}")

    p = write_photo_manifest(records, dry_run=dry_run)
    paths["photo_manifest"] = str(p)
    print(f"  3. {p}")

    p = write_missing_data_queue(records, dry_run=dry_run)
    paths["missing_data_queue"] = str(p)
    print(f"  4. {p}")

    summary["output_paths"] = paths
    summary["dry_run"] = dry_run
    summary["generated_ts"] = datetime.now(timezone.utc).isoformat()

    if dry_run:
        print("\n[DRY RUN] No files were written.")

    return summary


def main():
    parser = argparse.ArgumentParser(
        description="EXPORT-READY CONTROL LAYER v1 — compute readiness view and export artifacts"
    )
    parser.add_argument("--dry-run", action="store_true", help="Compute but don't write files")
    parser.add_argument("--pn", type=str, default=None, help="Process only this PN")
    args = parser.parse_args()

    summary = run(dry_run=args.dry_run, pn_filter=args.pn)
    print("\n" + json.dumps(
        {k: v for k, v in summary.items() if k not in ("coverage",)},
        indent=2, ensure_ascii=False
    ))


if __name__ == "__main__":
    main()
