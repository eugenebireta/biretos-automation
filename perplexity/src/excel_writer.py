# reserved for future Excel enrichment

from __future__ import annotations

from typing import Any, Awaitable, Callable, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Колонки по ТЗ:
# B  - Название товара или услуги
# F  - Описание
# R  - Изображения
# AO - Параметр: Модель
# AP - Параметр: Партномер
_NAME_COL = "B"
_DESC_COL = "F"
_IMAGE_COL = "R"
_MODEL_COL = "AO"
_PN_COL = "AP"


async def enrich_zapi_file(
    input_path: str,
    output_path: str,
    lookup_fn: Callable[[str, str], Awaitable[Dict[str, Any]]],
    brand: str = "ZAPI",
    pn_column: str = "A",
    start_row: int = 2,
) -> None:
    """
    Обогащает Excel-файл ZAPI данными из Perplexity и сохраняет в новый файл.

    - input_path: входной XLSX (например, ZAPI.xlsx)
    - output_path: выходной XLSX (например, ZAPI_enriched.xlsx)
    - lookup_fn: асинхронная функция (brand, pn) -> dict (результат lookup_product)
    - brand: бренд, по умолчанию 'ZAPI'
    - pn_column: буква колонки, где лежит исходный партномер (по умолчанию 'A')
    - start_row: с какой строки начинать обработку (1 — заголовки)
    """

    wb = load_workbook(input_path)
    ws = wb.active

    name_idx = column_index_from_string(_NAME_COL)
    desc_idx = column_index_from_string(_DESC_COL)
    image_idx = column_index_from_string(_IMAGE_COL)
    model_idx = column_index_from_string(_MODEL_COL)
    pn_idx = column_index_from_string(_PN_COL)
    pn_src_idx = column_index_from_string(pn_column)

    max_row = ws.max_row or 0

    for row in range(start_row, max_row + 1):
        raw_value = ws.cell(row=row, column=pn_src_idx).value
        pn = str(raw_value).strip() if raw_value is not None else ""
        if not pn:
            continue

        try:
            data = await lookup_fn(brand, pn)
        except Exception as exc:  # защита от любых неожиданных ошибок
            data = {
                "corrected_part_number": None,
                "model": None,
                "product_type": None,
                "image_url": None,
                "description": None,
                "raw": {"error": "excel_enrich_failed", "detail": str(exc)},
            }

        corrected_pn = (data.get("corrected_part_number") or pn).strip()
        model = (data.get("model") or "").strip() or None
        description = (data.get("description") or "").strip() or None
        image_url = (data.get("image_url") or "").strip() or None

        # Название товара или услуги: модель, если есть, иначе бренд + партномер
        if model:
            name = model
        else:
            name = f"{brand} {corrected_pn}"

        ws.cell(row=row, column=name_idx, value=name)
        ws.cell(row=row, column=desc_idx, value=description)
        ws.cell(row=row, column=image_idx, value=image_url)
        ws.cell(row=row, column=model_idx, value=model)
        ws.cell(row=row, column=pn_idx, value=corrected_pn)

    wb.save(output_path)


async def write_brand_samples_to_template(
    brand: str,
    count: int,
    input_path: str,
    output_path: str,
    sampler_fn: Callable[[str, int], Awaitable[List[Dict[str, Optional[str]]]]],
    start_row: int = 2,
) -> None:
    """
    Заполняет пустой шаблон Excel данными о товарах бренда, полученными из sampler_fn.
    """

    wb = load_workbook(input_path)
    ws = wb.active

    name_idx = column_index_from_string(_NAME_COL)
    desc_idx = column_index_from_string(_DESC_COL)
    image_idx = column_index_from_string(_IMAGE_COL)
    model_idx = column_index_from_string(_MODEL_COL)
    pn_idx = column_index_from_string(_PN_COL)

    _clear_columns(ws, start_row, [name_idx, desc_idx, image_idx, model_idx, pn_idx])

    samples = await sampler_fn(brand, count)
    if not samples:
        wb.save(output_path)
        return

    clean_brand = (brand or "").strip() or "Unknown"
    current_row = start_row

    for sample in samples:
        part_number = _sanitize_cell_value(sample.get("part_number"))
        model = _sanitize_cell_value(sample.get("model"))
        description = _sanitize_cell_value(sample.get("description"))
        image_url = _sanitize_cell_value(sample.get("image_url"))

        if not part_number and not model and not description:
            continue

        if model:
            name = model
        elif part_number:
            name = f"{clean_brand} {part_number}"
        else:
            name = clean_brand

        ws.cell(row=current_row, column=name_idx, value=name)
        ws.cell(row=current_row, column=desc_idx, value=description)
        ws.cell(row=current_row, column=image_idx, value=image_url)
        ws.cell(row=current_row, column=model_idx, value=model)
        ws.cell(row=current_row, column=pn_idx, value=part_number)

        current_row += 1

    wb.save(output_path)


def _clear_columns(ws, start_row: int, column_indexes: List[int]) -> None:
    """
    Очищает указанные колонки начиная со start_row до текущего конца таблицы.
    """

    max_row = ws.max_row or start_row
    for row in range(start_row, max_row + 1):
        for idx in column_indexes:
            ws.cell(row=row, column=idx, value=None)


def _sanitize_cell_value(value: Any) -> Optional[str]:
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed or None
    return None
